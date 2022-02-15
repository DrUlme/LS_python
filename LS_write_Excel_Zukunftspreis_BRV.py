import os, sys, sqlite3
from time import strftime
from time import gmtime
import time
# library für Excel:
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.styles import numbers
# from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.worksheet.defined_name import DefinedName
# image - using pillow?
from openpyxl.drawing.image import Image
#========================================================================
# classopenpyxl.workbook.protection.FileSharing


import LSglobal

BisZuJahr = LSglobal.RefJahr - 20
#========================================================================
# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect( LSglobal.SQLiteFile )

# Datensatzcursor erzeugen
cursor_R = connection.cursor()
cursor   = connection.cursor()
cursor_N = connection.cursor()
cursor_V = connection.cursor()
RBcursor = connection.cursor()
Pcursor  = connection.cursor()

#====================================================================================== Farben definieren
FillCol = "44ff44"
grayFill = PatternFill(start_color='666666',end_color='666666',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')

# Erstellen eines Workbooks:
wb = Workbook()
# wb = openpyxl.Workbook()
ws = wb.active
ws.title = "HLS_Erlangen_" + str(LSglobal.Jahr)

book = Workbook()
sheet = book.active

# ==============================================================================================================
ws.merge_cells('B1:H1')
ws['B1'] = LSglobal.Zeit + "-Langstrecke des Bayerischen Ruderverbandes in Erlangen " + LSglobal.Datum + " " + str(LSglobal.Jahr)
# " 24. Oktober 2020"
ws['B3'] = "1. Ergebnis"
#
#ws['H1'].alignment = Alignment(horizontal="center", vertical="bottom")

zeile = 5


# Laufende Nummer
LNR  = 0

# weibliche Athleten
NAW = 0

# männliche Athleten
NAM = 0

# Vereine mit teilnehmenden Athleten
NV  = 0

# VAnzahl der stattgefundenen Rennen
NR  = 0
# ________________ Ruderer über 6000 m



# Ruderverein
FillCol = "44ff44"
redFill = PatternFill(start_color='EE1111',end_color='EE1111',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')
# 11EE11 oder 44ff44


# ====================================================================== adapt column with
# indRe  = 'A' - indPos = 'B' - indSNr = 'C' - indStT = 'D' - indVor = 'E' - indNam = 'F' - indJah = 'G' - indEV  = 'H'- indCom = 'I'- indBot = 'J'

ws.column_dimensions['A'].width  = "1"

#indLNR = 'B'
#indRNR = 'C'
#indRe  = 'D'
#indMtr = 'E'
#indSNr = 'F'
#indNam = 'G'
#indGEN = 'H'
#indJah = 'I'
#
#indVN  = 'J'
#indEV  = 'K'
#
#indEZt = 'L'
#indBot = 'M'

# fixiere Tabelle:
ws.freeze_panes = ws['A5']

# erstelle Filter
# maxCols = str( zeile + 10 ) auf 256 gesetzt
ws.auto_filter.ref = "B4:G256"


# ______________________________________ set Logo
logo = Image("RVE_BRV_Flag.png")

# A bit of resizing 
logo.height = 58
logo.width = 158
#logo.height = 77
#logo.width = 210
#logo.width = 294
ws.merge_cells('O1:R3')
ws.add_image(logo, "O1")
ws['O1'].alignment = Alignment(horizontal="center",vertical="center")
# ws['O1'] = LSglobal.Name

ws['B4'] = 'lfd Nr.'	
ws['B4'].font = Font(name='arial', sz=11, b=True, i=False, color='222222')
 
ws['C4'] = 'Vorname'
ws['D4'] = 'Nachname' 	
ws['E4'] = 'Jahrgang'

ws['F4'] = 'Verein'
ws['G4'] = 'Punkte'



zeile = 4
lNR   = 0

# SQL-Abfrage
sql = "SELECT * FROM verein "
cursor_V.execute(sql)
for dsatz in cursor_V:
   print(dsatz[0] + ": start nach Zeile " + str(zeile + lNR))
   v_Athlet = 0
   # suche nach Ruderern
   sql = "SELECT * FROM ruderer WHERE verein = '" + dsatz[1] +"' and jahrgang > " + str(BisZuJahr) + " "
   cursor_R.execute(sql)
   #
   for ds in cursor_R:
      # ______________________________ suche nach den Booten pro Ruderer
      sql = "SELECT * FROM r2boot  WHERE rudererNr = " + str(ds[0]) 
      RBcursor.execute(sql)
      #
      for RudInd in RBcursor:   # for iR in range(0, (len(RudInd) - 2)):   
         # print(RudInd)
         sql = "SELECT * FROM boote WHERE nummer = " + str(RudInd[2]) + "  "
         cursor.execute(sql)
         Rd = cursor.fetchone()
         # check - nicht abgemeldet:
         if(Rd[10] == 0):
               v_Athlet = v_Athlet + 1
               lNR = lNR + 1
               if(lNR == 1):
                  ws['B' + str(zeile + lNR)] = '=1'
               else:
                  ws['B' + str(zeile + lNR)] = '=B' + str(zeile + lNR - 1) + '+ 1'
               ws['C' + str(zeile + lNR)] = ds[1]
               ws['D' + str(zeile + lNR)] = ds[2]
               ws['E' + str(zeile + lNR)] = str(ds[4])
               ws['F' + str(zeile + lNR)] = dsatz[0]
               ws['G' + str(zeile + lNR)] = '2'
               # ws['F' + str(zeile)].font = Font(name='arial', sz=11, b=True, i=False, color='222222')
      
# =================================== Data validation für Bootsklasse
connection.close()

# ______________________________________ save
wb.save('Zukunftspreis_H2021_BRV.xlsx')
