#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Oct 10 12:59:46 2021

@author: ulf

Aktuelle Kader-Liste (Stichtage: 01.01. und 01.07.) unter
<https://www.ruderverband.de/cms/home/leistungssport/kaderliste/Kader.xhtml>

"""
import os, sys, sqlite3
import time

# Excel
#========================================================================
from openpyxl import load_workbook

# globale Parameter
import LSglobal

Pfad_LS        = os.getcwd()
print("Starte aus Pfad '" + Pfad_LS + "'")

#========================================================================
# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect( LSglobal.SQLiteFile )

# Datensatzcursor erzeugen
cursor = connection.cursor()


if os.path.exists( "BRV_Kader.db" ):
    print("Datei bereits vorhanden")
    # os.remove("BRV_Kader.db")
    # sys.exit(0)
    KaderDB = sqlite3.connect("BRV_Kader.db" )
    cursorK = KaderDB.cursor()
    #
else:
   #========================================================================================== Einlesen der Datenbank
   # Verbindung zur Datenbank erzeugen
   KaderDB = sqlite3.connect( "BRV_Kader.db" )

   # Datensatzcursor erzeugen
   cursorK = KaderDB.cursor()

   # Tabellen erzeugen
   sql = "CREATE TABLE kader(" \
      "nummer INTEGER PRIMARY KEY, " \
      "name TEXT, " \
      "vorname TEXT, " \
      "verein TEXT, " \
      "jahrgang INTEGER, " \
      "kader TEXT, " \
      "foerder TEXT)"
   cursorK.execute(sql)
   # Add meta file for the creation date
   sql = "CREATE TABLE meta( name TEXT, wert TEXT)"
   cursorK.execute(sql)
   KaderDB.commit()
   
   DateString = time.strftime('%d.%m.%Y - %H:%M')
   sql = "INSERT INTO meta VALUES( 'creation', '" + DateString + "' )"
   cursorK.execute(sql)
   KaderDB.commit()
   #
   LNR = 0
   #
   # öffne excel
   wb = load_workbook( 'BRV_Kaderliste.xlsx' )
   # nutze das aktive Arbeitsblatt als 'ws'
   ws = wb.active
   #
   # loop über die Zeilen:
   iL = 1
   while iL < 104:
      # print(iL)
      # Jahrgang
      jahrgang = ws['F' + str(iL)].value
      
      if (jahrgang != None and isinstance(jahrgang, int)):
         LNR += 1
         #
         # Nummer 
         Nummer   = ws['B' + str(iL)].value
         # Vorname
         vorname  = str.strip(ws['D' + str(iL)].value)
         # Nachname
         name     = str.strip(ws['C' + str(iL)].value)
         # Verein
         verein   = str.strip(ws['E' + str(iL)].value)
         # Kader
         kader    = str.strip(ws['G' + str(iL)].value)
         # Förderstatus
         forder   = str.strip(ws['H' + str(iL)].value)
         #
         # print( str(Nummer) + ": " + vorname + " " + name + " " + str(jahrgang))
         sql = "INSERT INTO kader VALUES( " + str(LNR) + \
            ", '" + name + "', '" + vorname + "', '" + verein + "', " + str(jahrgang) + \
            ", '" + kader + "', '" + forder + "' )"
         cursorK.execute(sql)
         KaderDB.commit()
      #---
      iL += 1

#========================================================================================== suche nun nach Spielern
sqlite_select_query = """SELECT count(*) from kader"""
cursorK.execute(sqlite_select_query)
tmp = cursorK.fetchone()
nKader = tmp[0]
print("Anzahl der Ruderer in Datenbank:  ", nKader)

sql = "SELECT * FROM kader " # "WHERE nummer < 114"
cursorK.execute(sql)
for dsatz in cursorK:
      # Nachname
      name     = dsatz[1]
      vorname  = dsatz[2]
      # Verein
      verein   = dsatz[3]
      # Jahrgang
      jahrgang = dsatz[4]
      # Kader
      kader    = dsatz[5]
      # Förderstatus
      forder   = dsatz[6]
      #
      sql = "SELECT id FROM ruderer WHERE name='" + name + \
           "' and vorname='" + vorname + "' and jahrgang= " + str(jahrgang) 
      cursor.execute(sql)                
      RHelp= cursor.fetchone()
      if RHelp != None:
         # Ruderer[iP] = RHelp[0]
         sql = "UPDATE ruderer SET kader = '" + kader + " (" + forder + ")' WHERE id = '" + RHelp[0] + "'"
         cursor.execute(sql)
         connection.commit()
      else:
         print( vorname + " " + name + " (" + str(jahrgang) + " / " + kader + ") " + verein + " - nicht gestartet ?" )
         
connection.close()
KaderDB.close()