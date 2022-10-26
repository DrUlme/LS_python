import os, sys, sqlite3
from time import strftime
from time import gmtime
import time
# library für Excel:
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.styles import numbers
# from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.worksheet.defined_name import DefinedName
# image - using pillow?
from openpyxl.drawing.image import Image
#========================================================================
# classopenpyxl.workbook.protection.FileSharing


import LSglobal

#========================================================================
# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect( LSglobal.SQLiteFile )

# Datensatzcursor erzeugen
cursor_R = connection.cursor()
cursor   = connection.cursor()
cursor_N = connection.cursor()
cursor_V = connection.cursor()
RBcursor = connection.cursor()
Pcursor  = connection.cursor()

#====================================================================================== Farben definieren
FillCol = "44ff44"
grayFill = PatternFill(start_color='666666',end_color='666666',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')

# Erstellen eines Workbooks:
wb = Workbook()
# wb = openpyxl.Workbook()
ws = wb.active
ws.title = LSglobal.Zeit + "LS_Erlangen_" + str(LSglobal.Jahr)

book = Workbook()
sheet = book.active

# ==============================================================================================================
ws.merge_cells('B1:H1')
ws['B1'] = LSglobal.Zeit + "-Langstrecke des Bayerischen Ruderverbandes in Erlangen " + LSglobal.Datum + " " + str(LSglobal.Jahr)
# " 24. Oktober 2020"
ws['B3'] = "1. Ergebnis"
#
#ws['H1'].alignment = Alignment(horizontal="center", vertical="bottom")

zeile = 5

# in Zeile 5:
# lfd  | R_NR | R_Name | Strecke | Startnr | Ruderer (Vorname Name) | Geschlecht | Jahrgang | V_Nr | Verein | EndZeit | TNr | 

# Zusammenfassung ab O4 Verteilung 1: Anzahl Ruderer, männlich, weiblich, Vereine, Rennen
#                       Verteilung 2: Athleten pro Verein, Nummer, Verein (aufsteigend)
#                       Verteilung 3: Alters- und Gewichtsklassen: Junioren B, Lgw; ... Juniorinnen B ... JMA, JFA, SM A/B, SF A/B
indLNR = 'B'
indRNR = 'C'
indRe  = 'D'
indMtr = 'E'
indSNr = 'F'
indNam = 'G'
indGEN = 'H'
indJah = 'I'

indVN  = 'J'
indEV  = 'K'

indEZt = 'L'
indBot = 'M'

ws[indLNR + str(zeile)] = "LNR"
ws[indLNR + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indRNR + str(zeile)] = "RNR"
ws[indRNR + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
ws[indRe + str(zeile)] = "Rennen"
ws[indRe + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indMtr + str(zeile)] = "Strecke"
ws[indMtr + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indSNr + str(zeile)] = "SNR"
ws[indSNr + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indNam + str(zeile)] = "Name 1"
ws[indNam + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indGEN + str(zeile)] = "Geschlecht"
ws[indGEN + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indJah + str(zeile)] = "Jahrgang"
ws[indJah + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indVN + str(zeile)] = "VNR"
ws[indVN + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indEV + str(zeile)] = "Verein"
ws[indEV + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indEZt + str(zeile)] = "Endzeit"
ws[indEZt + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

# ws[indBot + str(zeile)] = "TNR"
# ws[indBot + str(zeile)].font = Font(name='arial', sz=6, b=False, i=False, color='ddddff')
ws[indBot + str(zeile)] = "kg"
ws[indBot + str(zeile)].font = Font(name='arial', sz=10, b=True, i=False, color='ddddff')


ws.column_dimensions["A"].alignment = Alignment(horizontal='center')
ws.column_dimensions["B"].alignment = Alignment(horizontal='center')
ws.column_dimensions["C"].alignment = Alignment(horizontal='center')

ws.column_dimensions["F"].alignment = Alignment(horizontal='center')
ws.column_dimensions["G"].alignment = Alignment(horizontal='center')

ws.column_dimensions["I"].alignment = Alignment(horizontal='center')


# Lauufende Nummer
LNR  = 0

# weibliche Athleten
NAW = 0

# männliche Athleten
NAM = 0

# Vereine mit teilnehmenden Athleten
NV  = 0

# VAnzahl der stattgefundenen Rennen
NR  = 0
# ________________ Ruderer über 6000 m

JMA = 0
JMA_Lgw = 0
JMB = 0
JMB_Lgw = 0
SM = 0
SM_Lgw = 0
MM = 0
MM_Lgw = 0

JFA = 0
JFA_Lgw = 0
JFB = 0
JFB_Lgw = 0
SF = 0
SF_Lgw = 0
MF = 0
Mf_Lgw = 0

# create a local named range (only valid for a specific sheet)
# sheetid = wb.sheetnames.index('Sheet')
# private_range = openpyxl.workbook.defined_name.DefinedName('privaterange', attr_text='Sheet!$A$6', localSheetId=sheetid)
# wb.defined_names.append(private_range)

zeile = zeile + 1
# Ausgabe des Ergebnisses
# ============================================================================================================
# SQL-Abfrage
sql = "SELECT * FROM rennen WHERE status >= 1  AND  strecke LIKE '%000 m' "

# Empfang des Ergebnisses
cursor_R.execute(sql)
for dsatz in cursor_R:
   Rennen = dsatz[0]
   ReStr  = str(Rennen)
   ReBez  = dsatz[1]
   Meter  = str(dsatz[4])
   Gender = dsatz[2]
   # zeile = zeile + 1
   
   # SQL-Abfrage
   sql = "SELECT * FROM boote WHERE rennen = " + ReStr + " and abgemeldet = 0  ORDER BY zeit, zeit3000, secstart, planstart "
   # Empfang des Ergebnisses
   cursor.execute(sql)
   

   # indRe  = 'A' - indPos = 'B' - indSNr = 'C' - indStT = 'D' - indVor = 'E' - indNam = 'F' - indJah = 'G' - indEV  = 'H'- indCom = 'I'- indBot = 'J'
   Platz = 0
   Anz   = 0
   Last  = 0
   for ds in cursor:
      Boot   = ds[0]
      StNr   = ds[1]
      #
      if(Anz == 0):
         NR = NR + 1
      LNR = LNR + 1
      #______________________________ Anzahl der Ruderer und ihre Nummern in der Datenbank
      sql = "SELECT rudererNr FROM r2boot  WHERE bootNr = " + str(Boot) 
      RBcursor.execute(sql)
      #
      for RudInd in RBcursor: # for iR in range(0, (len(RudInd) - 2)):         
         sql = "SELECT * FROM ruderer WHERE nummer = " + str(RudInd[0])
         Pcursor.execute(sql)
         Rd = Pcursor.fetchone()
         #
         # ------------------------------------------------------------------
         #
         ws[indLNR + str(zeile)] = str(LNR)
         ws[indLNR + str(zeile)].font = Font(name='arial', sz=10, b=False, i=False, color='aaaaaa')
         #
         if(Rd[6] > 10):
            ws[indBot + str(zeile)].number_format = '0.0'
            ws[indBot + str(zeile)] = Rd[6]
            # ws[indBot + str(zeile)] = ds[0]
         ws[indBot + str(zeile)].font = Font(name='arial', sz=10, b=False, i=False, color='aaaaaa')
         #
         ws[indRNR + str(zeile)] = ReStr
         ws[indRNR + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
         ws[indRe + str(zeile)] = ReBez
         ws[indRe + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
         #
         ws[indMtr + str(zeile)] = Meter + "m"
         ws[indMtr + str(zeile)].font = Font(name='arial', sz=11, b=True, i=False, color='4444dd')
         #
         ws[indGEN + str(zeile)] = Gender
         ws[indGEN + str(zeile)].font = Font(name='arial', sz=11, b=True, i=False, color='4444dd')
         #
         # StartNr
         ws[indSNr + str(zeile)] = ds[1]
         ws[indSNr + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
         #_______________________________________________________________________________________________________ Zeiten
         #
         ws[indEZt + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
         ws[indEZt + str(zeile)] = strftime("%M:%S", gmtime(ds[9]))
         # ds 11 => 9
         ws[indEZt + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='0000ff')
         ws[indEZt + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
         #_______________________________________________________________________________________________________ Zeiten
         #
         ws[indRNR + str(zeile)].font = Font(name='arial', sz=8, b=False, i=False, color='0000ff')
         ws[indRe + str(zeile)].font = Font(name='arial', sz=8, b=False, i=False, color='0000ff')
         ws[indSNr + str(zeile)].font = Font(name='arial', sz=10, b=False, i=False, color='222222')
         #         
         #_______________________________________________________________________________________________________ Zeiten
         #
         # Nachname
         ws[indNam + str(zeile)] = Rd[1] + " " + Rd[2]
         # Jahrgang
         ws[indJah + str(zeile)] = str( Rd[4] )
         #
         if(Gender == 'M'):
            NAM = NAM + 1
            if(Rd[4] == (LSglobal.RefJahr - 16) or Rd[4] == (LSglobal.RefJahr - 15)):
               if(Rd[6] > 1):
                  JMB_Lgw = JMB_Lgw + 1
               else:
                  JMB = JMB + 1
            elif(Rd[4] == (LSglobal.RefJahr - 18) or Rd[4] == (LSglobal.RefJahr - 17)):
               if(Rd[6] > 1):
                  JMA_Lgw = JMA_Lgw + 1
               else:
                  JMA = JMA + 1
            elif(Rd[4] > (LSglobal.RefJahr - 18) ):
               if(Rd[6] > 1):
                  SM_Lgw = SM_Lgw + 1
               else:
                  SM = SM + 1
         else:
            NAW = NAW + 1
            if(Rd[4] == (LSglobal.RefJahr - 16) or Rd[4] == (LSglobal.RefJahr - 15)):
               if(Rd[6] > 1):
                  JFB_Lgw = JFB_Lgw + 1
               else:
                  JFB = JFB + 1
            elif(Rd[4] == (LSglobal.RefJahr - 18) or Rd[4] == (LSglobal.RefJahr - 17)):
               if(Rd[6] > 1):
                  JFA_Lgw = JFA_Lgw + 1
               else:
                  JFA = JFA + 1
            elif(Rd[4] > (LSglobal.RefJahr - 18) ):
               if(Rd[6] > 1):
                  SF_Lgw = SF_Lgw + 1
               else:
                  SF = SF + 1
         # Verein - Kurzform
         sql = "SELECT * FROM verein WHERE kurz = '" + Rd[7] + "' "
         # Empfang des Ergebnisses
         cursor_V.execute(sql)
         Vp = cursor_V.fetchone()
         ws[indEV  + str(zeile)] =  Vp[0]
         ws[indEV  + str(zeile)].font = Font(name='arial', sz=10, b=False, i=False, color='222222')
         # ws[indVN  + str(zeile)] =  Vp[0] # Vereins-Nummer
         zeile = zeile + 1





#========================================================================

# Ruderverein
FillCol = "44ff44"
redFill = PatternFill(start_color='EE1111',end_color='EE1111',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')
# 11EE11 oder 44ff44


# ====================================================================== adapt column with
# indRe  = 'A' - indPos = 'B' - indSNr = 'C' - indStT = 'D' - indVor = 'E' - indNam = 'F' - indJah = 'G' - indEV  = 'H'- indCom = 'I'- indBot = 'J'

ws.column_dimensions['A'].width  = "1"
ws.column_dimensions[indLNR].width = "8"
ws.column_dimensions[indRNR].width = "8"
ws.column_dimensions[indRe].width  = "12"
ws.column_dimensions[indMtr].width = "8"

ws.column_dimensions[indSNr].width = "8"
ws.column_dimensions[indNam].width = "24"
ws.column_dimensions[indJah].width = "10"
ws.column_dimensions[indVN].width  = "8"    # Verein
ws.column_dimensions[indEV].width  = "28"    # Verein

ws.column_dimensions[indEZt].width = "10" 

ws.column_dimensions[indBot].width = "8"

ws.column_dimensions['N'].width = "4"
ws.column_dimensions['O'].width = "12"
ws.column_dimensions['P'].width = "16"
ws.column_dimensions['Q'].width = "6"
ws.column_dimensions['R'].width = "16"
ws.column_dimensions['S'].width = "6"

#indLNR = 'B'
#indRNR = 'C'
#indRe  = 'D'
#indMtr = 'E'
#indSNr = 'F'
#indNam = 'G'
#indGEN = 'H'
#indJah = 'I'
#
#indVN  = 'J'
#indEV  = 'K'
#
#indEZt = 'L'
#indBot = 'M'

# fixiere Tabelle:
ws.freeze_panes = ws['A6']

# erstelle Filter
# maxCols = str( zeile + 10 ) auf 256 gesetzt
ws.auto_filter.ref = "B5:M256"


# ______________________________________ set Logo
logo = Image("RVE_BRV_Flag.png")

# A bit of resizing 
logo.height = 58
logo.width = 158
#logo.height = 77
#logo.width = 210
#logo.width = 294
ws.merge_cells('O1:R3')
ws.add_image(logo, "O1")
ws['O1'].alignment = Alignment(horizontal="center",vertical="center")
# ws['O1'] = LSglobal.Name

ws['O4'] = 'Verteilung 1'
ws['O4'].font = Font(name='arial', sz=11, b=True, i=False, color='222222')
 
ws['P5'] = 'Anzahl Athleten'
ws['Q5'] = (NAM + NAW)

ws['P6'] = 'männlich'
ws['Q6'] = (NAM)

ws['P7'] = 'weiblich'
ws['Q7'] = (NAW)

ws['P8'] = 'Vereine'
ws['Q8'] = '15'

ws['P9'] = 'Rennen'
ws['Q9'] = (NR)


# ==================================================================
ws['O10'] = 'Verteilung 2'
ws['O10'].font = Font(name='arial', sz=11, b=True, i=False, color='222222')
 
ws['P10'] = 'Athleten/pro Verein'
ws['P10'].font = Font(name='arial', sz=11, b=True, i=False, color='222222')
 
ws['R10'] = 'Verein'
ws['R10'].font = Font(name='arial', sz=11, b=True, i=False, color='222222')

zeile = 10
# SQL-Abfrage
sql = "SELECT * FROM verein "
cursor_V.execute(sql)
for dsatz in cursor_V:
   v_Athlet = 0
   # suche nach Ruderern
   sql = "SELECT * FROM ruderer WHERE verein = '" + dsatz[1] +"'"
   cursor_R.execute(sql)
   #
   for ds in cursor_R:
      # ______________________________ suche nach den Booten pro Ruderer
      sql = "SELECT * FROM r2boot  WHERE rudererNr = " + str(ds[0]) 
      RBcursor.execute(sql)
      #
      for RudInd in RBcursor:   # for iR in range(0, (len(RudInd) - 2)):   
         # print(RudInd)
         sql = "SELECT * FROM boote WHERE nummer = " + str(RudInd[1]) + "  "
         cursor.execute(sql)
         Rd = cursor.fetchone()
         # check - nicht abgemeldet:
         if(Rd[10] == 0):
            v_Athlet = v_Athlet + 1
      #
   #
   if(v_Athlet > 0):
      zeile = zeile + 1
      ws['Q' + str(zeile)] = (v_Athlet)
      ws['Q' + str(zeile)].font = Font(name='arial', sz=11, b=True, i=False, color='222222')
      ws['R' + str(zeile)] = dsatz[0]
      ws['R' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')

ws['Q8'] = zeile - 10

# ==================================================================
zeile = zeile + 1
ws['O' + str(zeile)] = 'Verteilung 3'
ws['O' + str(zeile)].font = Font(name='arial', sz=11, b=True, i=False, color='222222')

zeile = zeile + 1
ws['P' + str(zeile)] = 'Junioren B'
ws['P' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['R' + str(zeile)] = 'Junioren B Lgw.'
ws['R' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['Q' + str(zeile)] = (JMB)
ws['Q' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['S' + str(zeile)] = (JMB_Lgw)
ws['S' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')

zeile = zeile + 1
ws['P' + str(zeile)] = 'Juniorinnen B'
ws['P' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['R' + str(zeile)] = 'Juniorinnen B Lgw.'
ws['R' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['Q' + str(zeile)] = (JFB)
ws['Q' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['S' + str(zeile)] = (JFB_Lgw)
ws['S' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')

zeile = zeile + 1
ws['P' + str(zeile)] = 'Junioren A'
ws['P' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['R' + str(zeile)] = 'Junioren A Lgw.'
ws['R' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['Q' + str(zeile)] = (JMA)
ws['Q' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['S' + str(zeile)] = (JMA_Lgw)
ws['S' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')

zeile = zeile + 1
ws['P' + str(zeile)] = 'Juniorinnen A'
ws['P' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['R' + str(zeile)] = 'Juniorinnen A Lgw.'
ws['R' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['Q' + str(zeile)] = (JFA)
ws['Q' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['S' + str(zeile)] = (JFA_Lgw)
ws['S' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')

zeile = zeile + 1
ws['P' + str(zeile)] = 'SF A/B'
ws['P' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['R' + str(zeile)] = 'SF A/B Lgw.'
ws['R' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['Q' + str(zeile)] = (SF)
ws['Q' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['S' + str(zeile)] = (SF_Lgw)
ws['S' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')

zeile = zeile + 1
ws['P' + str(zeile)] = 'SM A/B'
ws['P' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['R' + str(zeile)] = 'SM A/B Lgw.'
ws['R' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['Q' + str(zeile)] = (SM)
ws['Q' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')
ws['S' + str(zeile)] = (SM_Lgw)
ws['S' + str(zeile)].font = Font(name='arial', sz=11, b=False, i=False, color='222222')


# =================================== Data validation für Bootsklasse
connection.close()

# ______________________________________ save
wb.save('Endergebnis_' + LSglobal.ZeitK + str(LSglobal.Jahr) + '_BRV.xlsx')
