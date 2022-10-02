import os, sys, sqlite3
from time import strftime
from time import gmtime
import time
# library für Excel:
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.styles import numbers
# from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.worksheet.defined_name import DefinedName
# image - using pillow?
from openpyxl.drawing.image import Image
#========================================================================
# classopenpyxl.workbook.protection.FileSharing


import LSglobal

#========================================================================
# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect( LSglobal.SQLiteFile )

# Datensatzcursor erzeugen
cursor_R = connection.cursor()
cursor   = connection.cursor()
cursor_N = connection.cursor()
Qcursor  = connection.cursor()
Pcursor  = connection.cursor()

#====================================================================================== Farben definieren
FillCol = "44ff44"
grayFill = PatternFill(start_color='666666',end_color='666666',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')

# Erstellen eines Workbooks:
wb = Workbook()
# wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Endergebnis"

#ws2 = wb.create_sheet("Rennen") # insert at first position
#ws2.sheet_properties.tabColor = "1072BA"

# define Startzeit

book = Workbook()
sheet = book.active

# ==============================================================================================================
#ws.merge_cells('H1:J5')
#ws['H1'] = LSglobal.Name
#
#ws['H1'].alignment = Alignment(horizontal="center", vertical="bottom")

zeile = 1

indRe  = 'A'
indPos = 'B'
indSNr = 'C'
indStT = 'D'
indVor = 'E'
indNam = 'F'
indJah = 'G'
indEV  = 'H'
indGew = 'I'

indEZt = 'J'
indZ03 = 'K'
indZ36 = 'L'

indKdr = 'M'
indCom = 'N'
indBot = 'O'


ws[indRe + str(zeile)] = "Rennen"
ws[indRe + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indPos + str(zeile)] = "Position"
ws[indPos + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indSNr + str(zeile)] = "Start-Nr"
ws[indSNr + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indStT + str(zeile)] = "Start-Zeit"
ws[indStT + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indVor + str(zeile)] = "Vorname"
ws[indVor + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indNam + str(zeile)] = "Nachname"
ws[indNam + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indJah + str(zeile)] = "Jahrgang"
ws[indJah + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indEV + str(zeile)] = "Verein"
ws[indEV + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indGew + str(zeile)] = "[kg]"
ws[indGew + str(zeile)].font = Font(name='arial', sz=12, b=False, i=True, color='4444dd')

ws[indCom + str(zeile)] = "Kommentar"
ws[indCom + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indEZt + str(zeile)] = "Endzeit"
ws[indEZt + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
ws[indZ03 + str(zeile)] = "0-3000 m"
ws[indZ03 + str(zeile)].font = Font(name='arial', sz=12, b=False, i=False, color='4444dd')
ws[indZ36 + str(zeile)] = "3-6000 m"
ws[indZ36 + str(zeile)].font = Font(name='arial', sz=12, b=False, i=False, color='4444dd')

ws[indBot + str(zeile)] = "int.Bootnr."
ws[indBot + str(zeile)].font = Font(name='arial', sz=6, b=False, i=False, color='ddddff')


ws.column_dimensions["A"].alignment = Alignment(horizontal='center')
ws.column_dimensions["B"].alignment = Alignment(horizontal='center')
ws.column_dimensions["C"].alignment = Alignment(horizontal='center')

ws.column_dimensions["F"].alignment = Alignment(horizontal='center')
ws.column_dimensions["G"].alignment = Alignment(horizontal='center')

ws.column_dimensions["I"].alignment = Alignment(horizontal='center')

ws.column_dimensions[indKdr].alignment = Alignment(horizontal='center')


StNr = 1

# create a local named range (only valid for a specific sheet)
# sheetid = wb.sheetnames.index('Sheet')
# private_range = openpyxl.workbook.defined_name.DefinedName('privaterange', attr_text='Sheet!$A$6', localSheetId=sheetid)
# wb.defined_names.append(private_range)

# Ausgabe des Ergebnisses
# ============================================================================================================
# SQL-Abfrage
sql = "SELECT * FROM rennen WHERE status >= 1  AND  strecke LIKE '%000 m' "

# Empfang des Ergebnisses
cursor_R.execute(sql)
for dsatz in cursor_R:
   Rennen = dsatz[0]
   ReStr  = str(Rennen)
   zeile = zeile + 1
   
  
   # Renn-Nummer
   ws[indRe + str(zeile)] = Rennen
   ws[indRe + str(zeile)].fill = (grayFill)
   ws[indRe + str(zeile)].font = Font(name='arial', sz=14, b=True, i=False, color='ffffff')
   
   # interne Nummer
   ws[indPos + str(zeile)] = "0"
   ws[indPos + str(zeile)].font = Font(name='arial', sz=9, b=True, i=False, color='666666')
   ws[indPos + str(zeile)].fill = (grayFill)
   
   # 1. Startnummer
   ws[indSNr + str(zeile)] = "=StartNr_" + ReStr 
   ws[indSNr  + str(zeile)].font = Font(name='arial', sz=9, b=True, i=False, color='666666')
   ws[indSNr  + str(zeile)].fill = (grayFill)
   
   # 1. Startzeit
   #   ws[indStT + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
   #   ws[indStT + str(zeile)] = "=Zeit_" + ReStr 
   #   ws[indStT + str(zeile)].font = Font(name='arial', sz=10, b=False, i=False, color='ffffff')
   ws[indStT + str(zeile)].fill = (grayFill)
   #   ws[indStT + str(zeile)].alignment = Alignment(horizontal="left",vertical="center")
   if(dsatz[7] > 0):
      ws[indGew + str(zeile)] = str(dsatz[7])
      ws[indGew + str(zeile)].font = Font(name='arial', sz=10, b=False, i=True, color='ffffff')
      ws[indGew + str(zeile)].alignment = Alignment(horizontal='center', vertical='center')

   ws[indGew + str(zeile)].fill = (grayFill)
   ws[indEZt + str(zeile)].fill = (grayFill)
   ws[indZ03 + str(zeile)].fill = (grayFill)
   ws[indZ36 + str(zeile)].fill = (grayFill)
   
   # Renn-Bezeichnung
   #ws.merge_cells(indVor + str(zeile) + ':' + indJah + str(zeile))
   ws[indVor + str(zeile)] = dsatz[1]
   ws[indVor + str(zeile)].fill = (grayFill)
   ws[indVor + str(zeile)].font = Font(name='arial', sz=14, b=True, i=False, color='ffffff')
   
   # Streckenlänge
   #ws.merge_cells(indEV + str(zeile) + ':' + indCom + str(zeile))
   ws[indEV + str(zeile)] = dsatz[4]
   ws[indEV + str(zeile)].font = Font(name='arial', sz=14, b=True, i=False, color='ffffff')
   
   ws[indBot + str(zeile)] = 0
   ws[indBot + str(zeile)].fill = (grayFill)
   ws[indBot + str(zeile)].font = Font(name='arial', sz=6, b=False, i=False, color='666666')
   
   ws[indEV  + str(zeile)].fill = (grayFill)
   ws[indNam + str(zeile)].fill = (grayFill)
   ws[indJah + str(zeile)].fill = (grayFill)
   ws[indCom + str(zeile)].fill = (grayFill)
   ws[indKdr + str(zeile)].fill = (grayFill)
   
   # indRe  = 'A' - indPos = 'B' - indSNr = 'C' - indStT = 'D' - indVor = 'E' - indNam = 'F' - indJah = 'G' - indEV  = 'H'- indCom = 'I'- indBot = 'J'
   Platz = 0
   Anz   = 0
   Last  = 0
   #
   # SQL-Abfrage
   sql = "SELECT * FROM boote WHERE rennen = " + ReStr + " and abgemeldet = 0  ORDER BY zeit, zeit3000, secstart, planstart "
   # Empfang des Ergebnisses
   cursor.execute(sql)
   for ds in cursor:
      zeile = zeile + 1
      #
      ws[indRe + str(zeile)] = Rennen
      ws[indRe + str(zeile)].font = Font(name='arial', sz=8, b=False, i=False, color='0000ff')
      ws[indRe + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      # StartNr
      ws[indSNr + str(zeile)] = ds[1]
      ws[indSNr + str(zeile)].font = Font(name='arial', sz=10, b=False, i=False, color='222222')
      ws[indSNr + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      #_______________________________________________________________________________________________________ Zeiten
      ws[indStT + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
      ws[indStT + str(zeile)] = strftime("%H:%M:%S", gmtime(ds[4]))
      ws[indStT + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      #
      ws[indEZt + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
      ws[indEZt + str(zeile)] = strftime("%M:%S", gmtime(ds[9]))
      ws[indEZt + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='0000ff')
      ws[indEZt + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      #
      ws[indZ03 + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
      ws[indZ03 + str(zeile)] = strftime("%M:%S", gmtime(ds[7]))
      ws[indZ03 + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      #
      ws[indZ36 + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
      ws[indZ36 + str(zeile)] = strftime("%M:%S", gmtime(ds[8]))
      ws[indZ36 + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      #
      if(ds[9] == Last):
         Anz = Anz + 1
      else:
         Anz = Anz + 1
         Platz = Anz
         Last  = ds[9]
      # 
      ws[indPos + str(zeile)] = Platz
      # ws[indPos + str(zeile)].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")
      ws[indPos + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='0000ff')
      ws[indPos + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      #
      
      #Bemerkung
      ws[indCom + str(zeile)] = ds[11]
      ws[indCom + str(zeile)].alignment = Alignment(horizontal="left",vertical="center")
      
      # 
      ws[indBot + str(zeile)] = ds[0]
      ws[indBot + str(zeile)].font = Font(name='arial', sz=14, b=True, i=False, color='ffffff')
      
      sql = "SELECT * FROM r2boot  WHERE bootNr = " + str(ds[0]) 
      Qcursor.execute(sql)
      iP = 0
      for RudInd in Qcursor:         
         sql = "SELECT * FROM ruderer WHERE nummer = " + str(RudInd[3])
         Pcursor.execute(sql)
         Rd = Pcursor.fetchone()      
         #
         if(iP == 0):
            # Vorname
            Vorname = Rd[1]
            # Nachname
            Nachname = Rd[2]
            # Jahrgang
            Jahrgang = str( Rd[4] )
            # Verein
            Verein = Rd[7]
            # Gewicht
            if(Rd[6] <= 0):
               Gewicht = '-'
            else:
               Gewicht = str(Rd[6])
            Kader = Rd[8]
         else:
            # Vorname
            Vorname = Vorname + "\n" + Rd[1]
            # Nachname
            Nachname = Nachname + "\n" + Rd[2]
            # Jahrgang
            Jahrgang = Jahrgang + "\n" + str(Rd[4])
            # Verein
            Verein = Verein + "\n" + Rd[7]
            #Gewicht
            if(Rd[6] > 0):
               Gewicht = Gewicht + "\n" + str(Rd[6])
            elif(Gewicht != "-"):
               Gewicht = Gewicht + "\n-"
            Kader = Kader + "\n" + Rd[8]
         iP = iP + 1
      # Vorname
      ws[indVor + str(zeile)] = Vorname
      # Nachname
      ws[indNam + str(zeile)] = Nachname
      # Jahrgang
      ws[indJah + str(zeile)] = Jahrgang
      # Verein
      ws[indEV  + str(zeile)] = Verein
      # Gewicht
      ws[indGew + str(zeile)] = Gewicht
      # Kader
      ws[indKdr + str(zeile)] = Kader
      #
      if(iP > 1):
         ws.row_dimensions[ zeile ].height = 18*iP


# =================================== Data validation für Bootsklasse
connection.close()

#========================================================================

# Ruderverein
FillCol = "44ff44"
redFill = PatternFill(start_color='EE1111',end_color='EE1111',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')
# 11EE11 oder 44ff44


# ====================================================================== adapt column with
# indRe  = 'A' - indPos = 'B' - indSNr = 'C' - indStT = 'D' - indVor = 'E' - indNam = 'F' - indJah = 'G' - indEV  = 'H'- indCom = 'I'- indBot = 'J'

ws.column_dimensions[indRe].width  = "8"
ws.column_dimensions[indPos].width = "8"
ws.column_dimensions[indSNr].width = "8"
ws.column_dimensions[indStT].width = "11"
ws.column_dimensions[indVor].width = "14"
ws.column_dimensions[indNam].width = "14"
ws.column_dimensions[indJah].width = "10"
ws.column_dimensions[indEV].width  = "8"    # Verein
ws.column_dimensions[indGew].width = "5" 

ws.column_dimensions[indEZt].width = "10" 
ws.column_dimensions[indZ03].width = "8" 
ws.column_dimensions[indZ36].width = "8" 

ws.column_dimensions[indCom].width = "26"
ws.column_dimensions[indBot].width = "2"
ws.column_dimensions[indKdr].width = "8"

# fixiere Tabelle:
ws.freeze_panes = ws['A2']

# erstelle Filter
# maxCols = str( zeile + 10 ) auf 256 gesetzt
ws.auto_filter.ref = "A1:O256"


# ______________________________________ set Logo
logo = Image("RVE_BRV_Flag.png")

# A bit of resizing 
#logo.height = 77
#logo.width = 210
logo.height = 108
logo.width = 294
ws.merge_cells('P2:S7')
ws.add_image(logo, "P2")
ws.merge_cells('P1:S1')
ws['P1'].alignment = Alignment(horizontal="center",vertical="center")
ws['P1'] = LSglobal.Name
# ______________________________________ save
wb.save('Endergebnis_' +LSglobal.Zeit + str(LSglobal.Jahr) + '_test.xlsx')
