import os, sys, sqlite3
from time import strftime
from time import gmtime
import time
from openpyxl.styles import numbers

import LSglobal

# Excel
#========================================================================
# open existent workbook:
from openpyxl import load_workbook

filename = 'Startreihenfolge_offiziell.xlsx'
# filename = 'Startreihenfolge_H2020.xlsx'
# filename = LSglobal.StartXLS

# read the last used value (not the formula)
wb = load_workbook(filename, data_only=True)

print(wb.sheetnames)
# load worksheets: the last used one, may be wrong ...
#ws = wb.active
# Lese Meldungen
ws = wb["Meldungen"]
# ws = wb.get_sheet_by_name("Rennen")
#========================================================================
# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect( LSglobal.SQLiteFile )
#
# Datensatzcursor erzeugen
cursor  = connection.cursor()
cursorR = connection.cursor()
#
# --------------------------------------------------------------------------------------------------------------- Frühstarter
# hole Renn-Nummern für Früh und Spät-Starter
sql = "SELECT wert FROM meta WHERE name = 'Frühstarter'"
cursor.execute(sql)
Rd = cursor.fetchone()
Frühstart = Rd[0]

sql = "SELECT wert FROM meta WHERE name = 'Spätstarter'"
cursor.execute(sql)
Rd = cursor.fetchone()
Spätstart = Rd[0]

#========================================================================
zeile = 6

indRe  = 'A'
indFS  = 'B'
indPos = 'C'
indSNr = 'D'
indStT = 'E'
indVor = 'F'
indNam = 'G'
indJah = 'H'
indKdr = 'I'
indEV  = 'J'
indCom = 'K'
indLst = 'L'
indBot = 'N'
indHLP = 'M'

#
zeile = 7
while zeile > 6:
   #________________________________________________________________
   # integer Werte
   Rennen = ws[indRe + str(zeile)].value      # Rennen - Nummer
   FoderS = ws[indFS + str(zeile)].value      # 'F', 'S' oder '-'
   Positi = ws[indPos + str(zeile)].value     # Position im Startblock
   StNr   = ws[indSNr + str(zeile)].value     # vergebene Startnummer
   Boot   = ws[indBot + str(zeile)].value     # Boots-ID
   # Zeit-String
   sZeit  = str(ws[indStT + str(zeile)].value)
   #
   if(Boot == None):
      isBoot = 0
   elif( isinstance(Boot, int) ):
      if(Boot > 0):
         isBoot = 1
         Boot = str(Boot)
      else:
         isBoot = 0
   elif(len(Boot) > 0):
      isBoot = 1
   else:
      isBoot = 0
   #      print('---')
   #   else:
   if(isBoot > 0):
      #
      print("Zeile " + str(zeile) + ": _____________")
      # 
      Vornamen = ws[indVor + str(zeile)].value
      Vorname  = Vornamen.split("\n")
      #
      Namen    = ws[indNam + str(zeile)].value
      Name     = Namen.split("\n")
      #
      Jahrgang = str(ws[indJah + str(zeile)].value)
      Jahre    = Jahrgang.split("\n")
      #
      Vereine  = ws[indEV + str(zeile)].value
      Verein   = Vereine.split("\n")
      #
      print("Rennen " + str(Rennen) + "." + str(Positi) + "= #" + str(StNr) + " (" + Boot + ") :  " + sZeit)
      # setze das aktuelle Rennen als gesetzt (2)
      sql = "UPDATE rennen SET status = 2  WHERE nummer = " + str(Rennen)
      cursor.execute(sql)
      connection.commit()
      #
      #
      # SQL-Abfrage
      sql = "SELECT * FROM boote WHERE id = '" + Boot + "'"
      #
      # Empfang des Ergebnisses
      cursor.execute(sql)
      # for dsatz in cursor:
      dsatz = cursor.fetchone()
      # print(dsatz)
      # print("Rennen " + str(Rennen) + " = " + str(dsatz[2]) )
      # print("Verein " + Verein[0] + " = " + dsatz[3])
      # _______________________________________________________________________ korrekte Ruderer ?
      # _______________________________________________________________________ korrektes Rennen ?
      #
      if( Rennen != dsatz[3] ):
            print( "Boot " + Boot + " mit Startnr " + str(StNr) + " von Rennen " + str(dsatz[3]) + " nach " + str(Rennen) + " ?!" )
            # x = raw_input("Ändern? [Y/n]")
            x = input("Ändern? [Y/n] > ")
            if(x == "Y" or x== "y" or x == "j" or x == "J"):
               sql = "UPDATE boote SET rennen = " + str(Rennen) + " WHERE id = '" + Boot + "'"
               # print( sql )
               cursor.execute(sql)
               connection.commit()
               print("... geändert !")
      if( FoderS == "F" ):
         sql = "UPDATE boote SET alternativ = " + str(Frühstart) + " WHERE id = '" + Boot + "'"
         # print( sql )
         cursor.execute(sql)
         connection.commit()
      elif( FoderS == "S" ):
         sql = "UPDATE boote SET alternativ = " + str(Spätstart) + " WHERE id = '" + Boot + "'"
         # print( sql )
         cursor.execute(sql)
         connection.commit()
      # ______________________________________________________________________________________
      #
      sql = "UPDATE boote SET startnummer = " + str(StNr) + ", planstart = '" + sZeit + "' WHERE id = '" + Boot + "' "
      # print( sql )
      cursor.execute(sql)
      connection.commit()
      # dto. mit StNr
   #
   zeile = zeile + 1
   bZeit  = str(ws[indStT + str(zeile)].value)
   if(bZeit == None):
      zeile = 0
   if(len(bZeit) <= 5):
      zeile = 0
   #_____________________________________________________________


#______________________________________ EOF
