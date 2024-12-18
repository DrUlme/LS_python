# -*- coding: utf-8 -*-
"""
Schreibt die Meldungen in ein Excel-File, um damit einfachst die Startreihenfolge
zu setzen.

ToDo:
    Anpassen an neues Datenbankformat
    Schützen der Zellen, um Fehlbedienung zu vermeiden.
    (https://openpyxl.readthedocs.io/en/latest/protection.html)
    ws.protection.sheet = True
    ws.protection.password = '...'
    ws.protection.enable()      >>> ws.protection.disable()
    ws.protection.enable()

    ws.protection.sort = False
    ws.protection.selectUnlockedCells = True
    
    ws.auto_filter.ref = "A1:C2"
    ws.protection.autoFilter = False

    Zellen in Position festsetzen:  c = ws['B2']   &   ws.freeze_panes = c

"""

import os, sys, sqlite3
from time import strftime
from time import gmtime
import time
# library für Excel:
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.styles import numbers
# 
from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.worksheet.defined_name import DefinedName
# image - using pillow?
from openpyxl.drawing.image import Image
#========================================================================
# classopenpyxl.workbook.protection.FileSharing


import LSglobal

#========================================================================
# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect( LSglobal.SQLiteFile )

# Datensatzcursor erzeugen
cursor_R = connection.cursor()
cursor   = connection.cursor()
cursor_N = connection.cursor()
cursor_B = connection.cursor()
cursorRB = connection.cursor()
cursorRu = connection.cursor()

# hole Renn-Nummern für Früh und Spät-Starter
sql = "SELECT wert FROM meta WHERE name = 'Frühstarter'"
cursor.execute(sql)
Rd = cursor.fetchone()
Frühstart = int(Rd[0])
FrühStr = 'F'

sql = "SELECT wert FROM meta WHERE name = 'Spätstarter'"
cursor.execute(sql)
Rd = cursor.fetchone()
Spätstart = int(Rd[0])
SpätStr = 'S'

print("'alternativ' für Frühstart = " + str(Frühstart) + "   und Spätstart = " + str(Spätstart))

#======================================================================================
FillCol = "44ff44"
grayFill = PatternFill(start_color='666666',end_color='666666',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')

# Erstellen eines Workbooks:
wb = Workbook()
# wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Meldungen"

ws2 = wb.create_sheet("Rennen") # insert at first position
ws2.sheet_properties.tabColor = "1072BA"

# define Startzeit

book = Workbook()
sheet = book.active

# Create a data-validation object with list validation
dv = DataValidation(type="list", formula1='"F,S,-"', allow_blank=True)

sheet['A1'] = 56
sheet['A2'] = 43

# now = time.strftime("%b %d %Y %H:%M:%S", time.localtime(t))
ws['A3'].number_format = numbers.FORMAT_DATE_TIME4
ws['D3'].number_format = numbers.FORMAT_DATE_TIME4

#now = time.strftime(" %H:%M:%S", time.localtime(36000))
ws['A3'] = strftime("%H:%M:%S", gmtime(60*60*11))
# ws['A3']  = "10:00"
new_range = openpyxl.workbook.defined_name.DefinedName('Startzeit', attr_text='Meldungen!$A$3')
# wb.defined_names.append(new_range)   
wb.defined_names.add(new_range)   

#now = time.strftime(" %H:%M:%S", time.localtime(60))
ws['D3']  = strftime("%H:%M:%S", gmtime(60))
new_range = openpyxl.workbook.defined_name.DefinedName('Abstand', attr_text='Meldungen!$D$3')
wb.defined_names.add(new_range)
# wb.defined_names.append(new_range)   

# ============================================================================================================
# SQL-Abfrage
# sql = "SELECT * FROM rennen WHERE strecke != 'Athletik' and status > 0 ORDER BY nummer"
sql = "SELECT * FROM rennen WHERE strecke != 'Athletik'  ORDER BY nummer"
ws2['A1'] = "Nr."
ws2['B1'] = "Bezeichnung"
ws2['C1'] = "Boote"
ws2['D1'] = "frei"

ws2['E1'] = "1. Nummer"
ws2['F1'] = "1. Zeit"
ws2['G1'] = "Pause danach"


# Empfang des Ergebnisses
cursor_R.execute(sql)
zeile = 1
StartNr = 0
for dsatz in cursor_R:
   Rennen = dsatz[0]
   ReStr  = str(Rennen)
   zeile = zeile + 1
   ws2['A' + str(zeile)] = ReStr
   ws2['B' + str(zeile)] = dsatz[1]
   #   
   new_range = openpyxl.workbook.defined_name.DefinedName('Boote_' + ReStr, attr_text='Rennen!$C$' + str(zeile))
   wb.defined_names.add(new_range)
   # wb.defined_names.append(new_range)
   
   # Summe der Meldungen pro Rennen (ändert sich mit Änderung auf der Hauptseite
   ws2['C' + str(zeile)] = "=(SUMIF('Meldungen'!$M$7:$M$256,$A" + str(zeile) + " ) - $A"  + str(zeile) + ") / $A"  + str(zeile)
   
   # zusätzlich 2 Boote pro Rennen - wählbar
   ws2['D' + str(zeile)] = 2
   ws2['D' + str(zeile)].fill = (greenFill)
   #
   ws2['F' + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
   ws2['G' + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
   
   if(zeile < 3):
      ws2['E' + str(zeile)] = "1"
      ws2['F' + str(zeile)] = "=Startzeit"
   else:
      ws2['E' + str(zeile)] = "=$E" + str(zeile-1) + " + $C" + str(zeile-1) + " + $D" + str(zeile-1)
      ws2['F' + str(zeile)] = "=$F" + str(zeile-1) + " + ($C" + str(zeile-1) + " + $D" + str(zeile-1) + ")*Abstand + $G" + str(zeile-1)
   ws2['G' + str(zeile)] = "=1/24/60 * 0"
   new_range = openpyxl.workbook.defined_name.DefinedName('StartNr_' + ReStr, attr_text='Rennen!$E$' + str(zeile))
   wb.defined_names.add(new_range)
   # wb.defined_names.append(new_range)
   
   new_range = openpyxl.workbook.defined_name.DefinedName('Zeit_' + ReStr, attr_text='Rennen!$F$' + str(zeile))
   wb.defined_names.add(new_range)
   # wb.defined_names.append(new_range)   
# -----------------------------------------
ws2.column_dimensions['A'].width = "5"
ws2.column_dimensions['B'].width = "30"
ws2.column_dimensions['C'].width = "6"
ws2.column_dimensions['D'].width = "6"
ws2.column_dimensions['E'].width = "8"
ws2.column_dimensions['F'].width = "10"
ws2.column_dimensions['F'].width = "12"
ws2.column_dimensions["A"].alignment = Alignment(horizontal='center')
ws2.column_dimensions["C"].alignment = Alignment(horizontal='center')
ws2.column_dimensions["D"].alignment = Alignment(horizontal='center')
ws2.column_dimensions["E"].alignment = Alignment(horizontal='center')
ws2.column_dimensions["F"].alignment = Alignment(horizontal='center')
ws2.column_dimensions["G"].alignment = Alignment(horizontal='center')


# ==============================================================================================================
ws.merge_cells('H1:J5')
ws['H1'] = LSglobal.Name # "Langstrecke " + str(LSglobal.Jahr)

ws['H1'].alignment = Alignment(horizontal="center", vertical="bottom")

zeile = 6

indRe  = 'A'
indFS  = 'B'
indPos = 'C'
indSNr = 'D'
indStT = 'E'
indVor = 'F'
indNam = 'G'
indJah = 'H'
indKdr = 'I'
indEV  = 'J'
indCom = 'K'
indLst = 'L'
indBot = 'N'
indHLP = 'M'

ws[indRe + str(zeile)] = "Rennen"
ws[indRe + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indFS + str(zeile)] = "F/S"
ws[indFS + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indPos + str(zeile)] = "Position"
ws[indPos + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indSNr + str(zeile)] = "Start-Nr"
ws[indSNr + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
ws[indSNr + str(zeile)].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")

ws[indStT + str(zeile)] = "Start-Zeit"
ws[indStT + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
ws[indStT + str(zeile)].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")

ws[indVor + str(zeile)] = "Vorname"
ws[indVor + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indNam + str(zeile)] = "Nachname"
ws[indNam + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indJah + str(zeile)] = "Jahrgang"
ws[indJah + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indKdr + str(zeile)] = "Kader"
ws[indKdr + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indEV + str(zeile)] = "Verein"
ws[indEV + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indCom + str(zeile)] = "Bemerkung"
ws[indCom + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws[indLst + str(zeile)] = "letzte Zeit"
ws[indLst + str(zeile)].font = Font(name='arial', sz=6, b=False, i=False, color='4444dd')

ws[indBot + str(zeile)] = "int.Bootnr."
ws[indBot + str(zeile)].font = Font(name='arial', sz=6, b=False, i=False, color='4444dd')

ws[indHLP + str(zeile)] = "Start in"
ws[indHLP + str(zeile)].font = Font(name='arial', sz=6, b=False, i=False, color='4444dd')


ws.column_dimensions["A"].alignment = Alignment(horizontal='center')
ws.column_dimensions["B"].alignment = Alignment(horizontal='center')
ws.column_dimensions["C"].alignment = Alignment(horizontal='center')
ws.column_dimensions["D"].alignment = Alignment(horizontal='center')

ws.column_dimensions["G"].alignment = Alignment(horizontal='center')
ws.column_dimensions["H"].alignment = Alignment(horizontal='center')

ws.column_dimensions["K"].alignment = Alignment(horizontal='center')


StNr = 1

# create a local named range (only valid for a specific sheet)
# sheetid = wb.sheetnames.index('Sheet')
# private_range = openpyxl.workbook.defined_name.DefinedName('privaterange', attr_text='Sheet!$A$6', localSheetId=sheetid)
# wb.defined_names.append(private_range)

# Ausgabe des Ergebnisses
# ============================================================================================================
# SQL-Abfrage
sql = "SELECT * FROM rennen WHERE nummer < 114 AND boot IS NOT 'Athletik'"

# Empfang des Ergebnisses
cursor_R.execute(sql)
for dsatz in cursor_R:
   Rennen = dsatz[0]
   ReStr  = str(Rennen)
   
   RennBenamung = dsatz[3]
   Strecke      = dsatz[6]
   zeile = zeile + 1
   
   
   # Renn-Nummer
   ws[indRe + str(zeile)] = Rennen
   ws[indRe + str(zeile)].fill = (grayFill)
   ws[indRe + str(zeile)].font = Font(name='arial', sz=14, b=True, i=False, color='ffffff')
   
   # Früh-/Spätstarter
   ws[indFS + str(zeile)] = "-"
   ws[indFS + str(zeile)].fill = (grayFill)
   ws[indFS + str(zeile)].font = Font(name='arial', sz=14, b=True, i=False, color='666666')
   
   # interne Nummer
   ws[indPos + str(zeile)] = "0"
   ws[indPos + str(zeile)].font = Font(name='arial', sz=9, b=True, i=False, color='666666')
   ws[indPos + str(zeile)].fill = (grayFill)
   
   # 1. Startnummer
   ws[indSNr + str(zeile)] = "=StartNr_" + ReStr + " - 1e-8"
   ws[indSNr  + str(zeile)].font = Font(name='arial', sz=9, b=True, i=False, color='666666')
   ws[indSNr  + str(zeile)].fill = (grayFill)
   
   # 1. Startzeit
   ws[indStT + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
   ws[indStT + str(zeile)] = "=Zeit_" + ReStr + " - 1e-16"
   ws[indStT + str(zeile)].font = Font(name='arial', sz=10, b=False, i=False, color='ffffff')
   ws[indStT + str(zeile)].fill = (grayFill)
   ws[indStT + str(zeile)].alignment = Alignment(horizontal="left",vertical="center")
   
   # Renn-Bezeichnung
   #ws.merge_cells(indVor + str(zeile) + ':' + indJah + str(zeile))
   ws[indVor + str(zeile)] = RennBenamung
   ws[indVor + str(zeile)].fill = (grayFill)
   ws[indVor + str(zeile)].font = Font(name='arial', sz=14, b=True, i=False, color='ffffff')
   
   # Streckenlänge
   #ws.merge_cells(indEV + str(zeile) + ':' + indCom + str(zeile))
   ws[indEV + str(zeile)] = Strecke
   ws[indEV + str(zeile)].fill = (grayFill)
   ws[indEV + str(zeile)].font = Font(name='arial', sz=14, b=True, i=False, color='ffffff')
   
   ws[indBot + str(zeile)] = 0
   ws[indBot + str(zeile)].fill = (grayFill)
   ws[indBot + str(zeile)].font = Font(name='arial', sz=14, b=True, i=False, color='666666')
   
   ws[indKdr + str(zeile)].fill = (grayFill)
   ws[indNam + str(zeile)].fill = (grayFill)
   ws[indJah + str(zeile)].fill = (grayFill)
   ws[indCom + str(zeile)].fill = (grayFill)
   ws[indLst + str(zeile)].fill = (grayFill)
   ws[indBot + str(zeile)].fill = (grayFill)
   ws[indHLP + str(zeile)].fill = (grayFill)
   
   # indRe  = 'A' - indPos = 'B' - indSNr = 'C' - indStT = 'D' - indVor = 'E' - indNam = 'F' - indJah = 'G' - indEV  = 'H'- indCom = 'I'- indBot = 'J'
   
   # SQL-Abfrage
   sql = "SELECT * FROM boote WHERE rennen = " + ReStr + " and abgemeldet = 0"
   # Empfang des Ergebnisses
   cursor_B.execute(sql)

   for ds in cursor_B:
      zeile = zeile + 1
      BootIdx = ds[0]
      Altern  = ds[12]
      #______________________________ Anzahl der Ruderer und ihre Nummern in der Datenbank
      
      
      ws[indRe + str(zeile)] = Rennen
      ws[indRe + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='0000ff')
      ws[indRe + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      
      # print(str(Altern) + " == " + str(Frühstart) + " ?")
      if(Altern == Frühstart):
         ws[indFS + str(zeile)] = "F"
      elif(Altern == Spätstart):
         ws[indFS + str(zeile)] = "S"
      else:
         ws[indFS + str(zeile)] = "-"
      # dv.add(ws[indFS + str(zeile)])
      # in Excel: Daten => Gültigkeit => Zulassen:Liste => -\nF\nS
      ws[indFS + str(zeile)].fill = (greenFill)
      ws[indFS + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='0000ff')
      ws[indFS + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      
      # Startnummer und Zeit
      # ws[indSNr + str(zeile)] = "=INDIRECT(\"StartNr_\"& $A" + str(zeile) + ") - 1 + $B" + str(zeile)
      ws[indSNr + str(zeile)] = "=INDIRECT(\"StartNr_\"& $M" + str(zeile) + ") - 1 + $C" + str(zeile)
      ws[indSNr + str(zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='0000ff')
      ws[indSNr + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      
      # Zeit
      ws[indStT + str(zeile)].number_format = numbers.FORMAT_DATE_TIME4
      # ws[indStT + str(zeile)] = "=INDIRECT(\"Zeit_\"& $A" + str(zeile) + ") + ($C" + str(zeile) + " - 1)*Abstand"
      ws[indStT + str(zeile)] = "=INDIRECT(\"Zeit_\"& $M" + str(zeile) + ") + ($C" + str(zeile) + " - 1)*Abstand"
      ws[indStT + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      
      #Bemerkung
      ws[indCom + str(zeile)] = ds[13]
      ws[indCom + str(zeile)].alignment = Alignment(horizontal="left",vertical="center")
      
      # 
      ws[indPos + str(zeile)] = "1"
      ws[indPos + str(zeile)].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")
      ws[indPos + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      # 
      ws[indBot + str(zeile)] = ds[0]
      ws[indBot + str(zeile)].font = Font(name='arial', sz=10, b=True, i=False, color='cccccc')
      ws[indBot + str(zeile)].alignment = Alignment(horizontal="right",vertical="center")
      
      ws[indHLP + str(zeile)] = "=IF($B" + str(zeile) + "=\"F\", " + str(Frühstart) + ", IF($B" + str(zeile) + "=\"S\", " + str(Spätstart) + ", $A" + str(zeile) + "))"
      ws[indHLP + str(zeile)].font = Font(name='arial', sz=10, b=True, i=False, color='999999')
      ws[indHLP + str(zeile)].alignment = Alignment(horizontal="center",vertical="center")
      
      # suche nach Einträgen in r2boot for dieses Boot
      sql = "SELECT * FROM r2boot WHERE bootid = " + str(BootIdx)
      cursorRB.execute(sql)
      # loop über alle Einträge
      iP = 0
      for rbs in cursorRB:
         # suche nach Einträgen in r2boot for dieses Boot
         sql = "SELECT * FROM ruderer WHERE id = " + str(rbs[2])
         cursorRu.execute(sql)
         Rd = cursorRu.fetchone()
         if(iP == 0):
            # Vorname
            Vorname = Rd[1]
            # Nachname
            Nachname = Rd[2]
            # Jahrgang
            Jahrgang = str( Rd[4] )
            # Verein
            Verein = Rd[7]
            # Kader
            Kader = Rd[8]
         else:
            # Vorname
            Vorname = Vorname + "\n" + Rd[1]
            # Nachname
            Nachname = Nachname + "\n" + Rd[2]
            # Jahrgang
            Jahrgang = Jahrgang + "\n" + str(Rd[4])
            # Verein
            Verein = Verein + "\n" + Rd[7]
            # Kader
            Kader = Kader + "\n" + Rd[8]
         #
         iP += 1
      # Vorname
      ws[indVor + str(zeile)] = Vorname
      # Nachname
      ws[indNam + str(zeile)] = Nachname
      # Jahrgang
      ws[indJah + str(zeile)] = Jahrgang
      # Verein
      ws[indEV + str(zeile)] = Verein
      # Kader
      ws[indKdr + str(zeile)] = Kader
      #
      if(iP > 1):
         ws.row_dimensions[ zeile ].height = 18*iP



# =================================== Data validation für Bootsklasse
connection.close()

#========================================================================
ws['A1'] = "Bitte nur die grünen Zellen bearbeiten"
ws['A1'].font = Font(name='arial', sz=12, b=True, i=True, color='4444dd')

# Ruderverein
FillCol = "44ff44"
redFill = PatternFill(start_color='EE1111',end_color='EE1111',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')
# 11EE11 oder 44ff44

ws['A2'] = "Erste Startzeit"
ws['A2'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
ws.merge_cells('A3:B3')
ws['A3'].fill = PatternFill(start_color=FillCol, end_color="4444dd",  fill_type = "solid")

ws['D2'] = "Startabstand"
ws['D2'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
ws.merge_cells('D3:E3')
ws['D3'].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")


# ====================================================================== adapt column with
# indRe  = 'A' - indPos = 'B' - indSNr = 'C' - indStT = 'D' - indVor = 'E' - indNam = 'F' - indJah = 'G' - indEV  = 'H'- indCom = 'I'- indBot = 'J'

ws.column_dimensions[indRe].width  = "8"
ws.column_dimensions[indFS].width  = "6"
ws.column_dimensions[indPos].width = "8"
ws.column_dimensions[indSNr].width = "8"
ws.column_dimensions[indStT].width = "11"
ws.column_dimensions[indVor].width = "14"
ws.column_dimensions[indNam].width = "14"
ws.column_dimensions[indJah].width = "10"
ws.column_dimensions[indKdr].width = "8"
ws.column_dimensions[indEV].width  = "8"
ws.column_dimensions[indCom].width = "26"
ws.column_dimensions[indHLP].width = "8"
ws.column_dimensions[indBot].width = "8"

# fixiere Tabelle:
ws.freeze_panes = ws['A7']

# erstelle Filter
# maxCols = str( zeile + 10 ) auf 256 gesetzt
ws.auto_filter.ref = "A6:N256"

# ================================================= Sortieren
ws['D5'] = "hiermit sortieren"
ws['D5'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
ws.merge_cells('D5:E5')
ws['D5'].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")

# data valudation
ws.add_data_validation(dv)
dv.add("B7:B256")

# ______________________________________ set Logo
logo = Image("RVE_BRV_Flag.png")

# A bit of resizing 
logo.height = 77
logo.width = 210

ws.add_image(logo, "H1")

# ______________________________________ save
wb.save('Startreihenfolge_' + LSglobal.ZeitK + str(LSglobal.Jahr) + '.xlsx')
