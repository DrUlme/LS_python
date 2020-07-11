import os, sys, sqlite3

# library für Excel:
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.worksheet.datavalidation import DataValidation

#========================================================================


# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect("LS2020H.db")

# Datensatzcursor erzeugen
cursor = connection.cursor()

# Erstellen eines Workbooks:
wb = Workbook()
ws = wb.active
ws.title = "Meldebogen"

ws.column_dimensions["K"].alignment = Alignment(horizontal='center')
ws.column_dimensions["M"].alignment = Alignment(horizontal='center')
ws.column_dimensions["N"].alignment = Alignment(horizontal='center')
ws.column_dimensions["O"].alignment = Alignment(horizontal='center')
ws.column_dimensions["P"].alignment = Alignment(horizontal='center')

ws2 = wb.create_sheet("Rennen") # insert at first position
ws2.sheet_properties.tabColor = "1072BA"

# SQL-Abfrage
sql = "SELECT * FROM rennen WHERE nummer < 114"

# Empfang des Ergebnisses
cursor.execute(sql)

# Ausgabe des Ergebnisses
for dsatz in cursor:
   Rennen = dsatz[0]
   ws2['A' + str(Rennen)] = dsatz[0]
   ws2['B' + str(Rennen)] = dsatz[1]
   ws2['C' + str(Rennen)] = dsatz[2]
   ws2['D' + str(Rennen)] = dsatz[3]
   ws2['E' + str(Rennen)] = dsatz[4]
   ws2['F' + str(Rennen)] = dsatz[7]
   ws2['G' + str(Rennen)] = dsatz[8]
   ws2['H' + str(Rennen)] = dsatz[9]
   
   ws['K' + str(Rennen + 1)] = dsatz[0]
   ws['L' + str(Rennen + 1)] = dsatz[1]
   ws['M' + str(Rennen + 1)] = dsatz[4]
   ws['N' + str(Rennen + 1)] = dsatz[7]
   ws['O' + str(Rennen + 1)] = dsatz[8]
   ws['P' + str(Rennen + 1)] = dsatz[9]
   if( dsatz[7] < 0 ):
      ws2['F' + str(Rennen)] = ""
      ws['N' + str(Rennen + 1)] = "ohne"
   if( dsatz[3] == "all"):
      ws2['D' + str(Rennen)] = "-"
   ws['K' + str(Rennen + 1)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['L' + str(Rennen + 1)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['M' + str(Rennen + 1)].font = Font(name='arial', sz=10, b=False, i=False, color='4444dd')
   ws['N' + str(Rennen + 1)].font = Font(name='arial', sz=10, b=False, i=False, color='4444dd')
   ws['O' + str(Rennen + 1)].font = Font(name='arial', sz=10, b=False, i=False, color='4444dd')
   ws['P' + str(Rennen + 1)].font = Font(name='arial', sz=10, b=False, i=False, color='4444dd')
   ws['K' + str(Rennen + 1)].alignment = Alignment(horizontal="center")
   ws['M' + str(Rennen + 1)].alignment = Alignment(horizontal="center")
   ws['N' + str(Rennen + 1)].alignment = Alignment(horizontal="center")
   ws['O' + str(Rennen + 1)].alignment = Alignment(horizontal="center")
   ws['P' + str(Rennen + 1)].alignment = Alignment(horizontal="center")

# =================================== Data validation für Bootsklasse
dv = DataValidation(type="list", formula1='"-","1x","2-","2x","4x","4x+"', allow_blank=True)
ws.add_data_validation(dv)

#dv.add(ws["H"+str(5)])
# ws2['K1'] = "Bootsklassen"
# ws2['K2'] = "-"
# ws2['K3'] = "1x"
# ws2['K4'] = "2-"
# ws2['K5'] = "2x"
# ws2['K6'] = "4x"
# ws2['K7'] = "4-"
# ws2['K8'] = "4x+"


print(dsatz[0], dsatz[1], dsatz[2],
          dsatz[3], dsatz[4], dsatz[5])
ws['K1'] = "Nr"
ws['L1'] = "Rennen"
ws['M1'] = "Strecke"
ws['N1'] = "kg"
#          Jahrgang
ws['O1'] = "von"
ws['P1'] = "bis"

# Verbindung beenden
connection.close()

#========================================================================
ws['A1'] = "Bitte nur die grünen Zellen bearbeiten"
ws['A1'].font = Font(name='arial', sz=12, b=True, i=True, color='4444dd')

# Ruderverein
FillCol = "44ff44"
redFill = PatternFill(start_color='EE1111',end_color='EE1111',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')
# 11EE11 oder 44ff44

ws['A2'] = "Langform Verein"
ws['A2'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
ws.merge_cells('A3:E3')
# ws['A3'].fill = Style("Good")
ws['A3'].fill = PatternFill(start_color=FillCol, end_color="4444dd",  fill_type = "solid")

ws['G2'] = "Kurzform Verein"
ws['G2'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
ws.merge_cells('G3:H3')
ws['G3'].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")

ws['A4'] = "Adresse Verein"
ws['A4'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
ws.merge_cells('A5:E5')
ws['A5'].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")
ws.merge_cells('A6:E6')
ws['A6'].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")

# Betreuer
ws['A7'] = "Betreuer: Vorname"
ws['A7'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
ws.merge_cells('A8:B8')
ws['A8'].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")

ws['C7'] = "Nachname"
ws['C7'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
#ws.merge_cells('D8:D8')
ws['C8'].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")

ws['E7'] = "Telefon"
ws['E7'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
ws.merge_cells('E8:F8')
ws['E8'].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")

ws['H7'] = "e-mail"
ws['H7'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')
ws.merge_cells('H8:I8')
ws['H8'].fill = PatternFill(start_color=FillCol, end_color=FillCol,  fill_type = "solid")


ws['A10'] = "Meldung"
ws['A10'].font = Font(name='arial', sz=8, b=True, i=False, color='4444dd')

ws['B10'] = "Renn-Nr"
ws['B10'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws['C10'] = "Vorname"
ws['C10'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws['D10'] = "Nachname"
ws['D10'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws['E10'] = "Jahrgang"
ws['E10'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws['F10'] = "Rgm."
ws['F10'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws['G10'] = "Rennen"
ws['G10'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
ws['H9'] = "Boots"
ws['H9'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
ws['H10'] = "Klasse"
ws['H10'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

ws.merge_cells('I10:J10')
ws['I10'] = "Kommentare"
ws['I10'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

# erstes Rennen
ws['A11'] = 1
ws['B11'] = 4

#
# ws.conditional_formatting.add('C11:D14',FormulaRule(formula=['$A11>0'], stopIfTrue=True, fill=greenFill))

# 
for ROW in range(11,54):
   ws.merge_cells('I'+ str(ROW) + ':J' + str(ROW))
   # ws.merge_cells('F'+ str(ROW) + ':G' + str(ROW))
   ws['G'+ str(ROW)].alignment = Alignment(horizontal="right", shrinkToFit=True)
   ws['G' + str(ROW)] = "=IF(ISNUMBER($B" + str(ROW) + "),INDIRECT(\"Rennen!$B\"&($B" + str(ROW) + ")),\"\")"
   ws['H' + str(ROW)] = '=IF(ISNUMBER($B' + str(ROW) + '),INDIRECT("Rennen!$D"&($B' + str(ROW) + ')),"")'
   dv.add(ws["H"+str(ROW)])
   ws.conditional_formatting.add('H' + str(ROW),FormulaRule(formula=['$H' + str(ROW) + '="-"' ], stopIfTrue=True, fill=greenFill))
   # B & I gleiche Formatierung:
   ws.conditional_formatting.add('B' + str(ROW),FormulaRule(formula=['ROUND($A' + str(ROW) + ')=$A' +str(ROW)], stopIfTrue=True, fill=greenFill))
   ws.conditional_formatting.add('I' + str(ROW),FormulaRule(formula=['ROUND($A' + str(ROW) + ')=$A' +str(ROW)], stopIfTrue=True, fill=greenFill))
   myRange = 'C' + str(ROW) + ':D' + str(ROW)
   myForml = '$A' + str(ROW) + '>0'
   ws.conditional_formatting.add(myRange,FormulaRule(formula=['ISNUMBER($A' + str(ROW) + ')' ], stopIfTrue=True, fill=greenFill))
   ws.conditional_formatting.add(myRange,FormulaRule(formula=['$A' + str(ROW) + '="-"' ], stopIfTrue=True, fill=greenFill))
   ws.conditional_formatting.add('F' + str(ROW),FormulaRule(formula=['ISNUMBER($A' + str(ROW) + ')' ], stopIfTrue=True, fill=greenFill))
   ws.conditional_formatting.add('F' + str(ROW),FormulaRule(formula=['$A' + str(ROW) + '="-"' ], stopIfTrue=True, fill=greenFill))
   #
   # ======================================================================  Check Jahrgang:
   # Hilfe für Rennen
   ws['R' + str(ROW)] = '=IF(ISBLANK($A' + str(ROW) + '),0,IF(ISNUMBER($B' + str(ROW) + '),$B' + str(ROW) + ',IF($A' + str(ROW) + '="-",IF(ISNUMBER($B' + str(ROW-1) + '),$B' + str(ROW-1) \
   + ',IF(ISNUMBER($B' + str(ROW-2) + '),$B' + str(ROW-2) + ',IF(ISNUMBER($B' + str(ROW-3) + '),$B' + str(ROW-3) + ', 0))))))'
   # Check des Alters
   ws['Q' + str(ROW)] = '=IF(ISBLANK($A' + str(ROW) + '),0,IF($R' + str(ROW) + '<1,0,IF($E' + str(ROW) + '<1,1,IF($E' + str(ROW) + \
   '<INDIRECT("Rennen!$G"&($R' + str(ROW) + ')),2,IF($E' + str(ROW) + '>INDIRECT("Rennen!$H"&($R' + str(ROW) + ')),2,1)))))'
   # ws.conditional_formatting.add('I' + str(ROW) ,FormulaRule(formula=['$B' + str(ROW) + '>0'], stopIfTrue=True, fill=greenFill))
   # Anpassen der Farbe für Jahrgang analog des Check-Ergebnisses
   ws.conditional_formatting.add('E' + str(ROW) ,FormulaRule(formula=['$Q' + str(ROW) + '=2'], stopIfTrue=True, fill=redFill))
   ws.conditional_formatting.add('E' + str(ROW) ,FormulaRule(formula=['$Q' + str(ROW) + '=1'], stopIfTrue=True, fill=greenFill))
   #
   #___________________________ interne Nummer - oder Meldung - Spalte A
   if(ROW > 11):
      ws['A' + str(ROW)] = '= IF(H' + str(ROW-1) + '="1x",A' + str(ROW-1) + '+1,IF(H' + str(ROW-2) + '="2-",A' + str(ROW-2) \
      + '+1,IF(H' + str(ROW-2) + '="2x",A' + str(ROW-2) + '+1,IF(H' + str(ROW-4) + '="4x",A' + str(ROW-4) + '+1,IF(H' + str(ROW-4) \
      + '="4x+","Stm.",IF(H' + str(ROW-5) + '="4x+",A' + str(ROW-5) + '+1,IF(H' + str(ROW-1) + '="2-","-","")))))))'



# ====================================================================== adapt column with
ws.column_dimensions['A'].width = "6"
ws.column_dimensions['B'].width = "8"
ws.column_dimensions['C'].width = "16"
ws.column_dimensions['D'].width = "16"
ws.column_dimensions['E'].width = "10"
ws.column_dimensions['F'].width = "8"
ws.column_dimensions['G'].width = "14"
ws.column_dimensions['H'].width = "8"
ws.column_dimensions['I'].width = "12"
ws.column_dimensions['J'].width = "10"
ws.column_dimensions['K'].width = "5"
ws.column_dimensions['L'].width = "20"
ws.column_dimensions['M'].width = "8"
ws.column_dimensions['N'].width = "6"
ws.column_dimensions['O'].width = "7"
ws.column_dimensions['P'].width = "7"
# base alignment
ws.column_dimensions["A"].alignment = Alignment(horizontal='left')
ws.column_dimensions["B"].alignment = Alignment(horizontal='center')
ws.column_dimensions["E"].alignment = Alignment(horizontal='center')
ws.column_dimensions["H"].alignment = Alignment(horizontal='center')
# ws.column_dimensions["K"].alignment = Alignment(horizontal='center')
# ws.column_dimensions["M"].alignment = Alignment(horizontal='center')
# ws.column_dimensions["N"].alignment = Alignment(horizontal='center')
# ws.column_dimensions["O"].alignment = Alignment(horizontal='center')
# ws.column_dimensions["P"].alignment = Alignment(horizontal='center')

# fixiere Tabelle:
ws.freeze_panes = ws['C11']

# setze Ausrichtung für Telefon-Nummer auf links zurück
ws['E8'].alignment = Alignment(horizontal="left")

wb.save('Meldebogen_H2020.xlsx')
