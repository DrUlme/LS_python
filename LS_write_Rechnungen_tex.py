#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Oct  8 15:48:30 2020

@author: ulf
"""
import time
import math
import os, sys, sqlite3

# library für Excel:
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.styles import numbers
# from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.worksheet.defined_name import DefinedName
# image - using pillow?

import LSglobal

#========================================================================
# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect( LSglobal.SQLiteFile )
#
# Datensatzcursor erzeugen
Rcursor  = connection.cursor()
Bcursor  = connection.cursor()
Pcursor  = connection.cursor()
Qcursor  = connection.cursor()
Vcursor  = connection.cursor()
#
cursor  = connection.cursor()
#
#====================================================================================== Farben definieren
FillCol = "44ff44"
grayFill = PatternFill(start_color='666666',end_color='666666',fill_type='solid')
greenFill = PatternFill(start_color='44ff44',end_color='44ff44',fill_type='solid')
noFill = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')

# Erstellen eines Workbooks:
wb = Workbook()
# wb = openpyxl.Workbook()
ws = wb.active
# ws.title = LSglobal.Name
ws.title = "LS" + str(LSglobal.Jahr) + LSglobal.ZeitK

book = Workbook()
sheet = book.active

# ==============================================================================================================
ws.merge_cells('B1:H1')
ws['B1'] = LSglobal.Name
ws['B3'] = "Rechnungen"
ws['A4'] = "Verein"
ws['B4'] = "Kürzel"
ws['C4'] = "Betrag"
ws['D4'] = "e-mails"

wsNr = 5

t1 = time.localtime()

TXT_1 = "\\documentclass[DIN,\n	fromalign=right,\n	fromurl=on,\n	fromemail=on,\n	fromrule=on,\n	fromlogo=on\n]{scrlttr2}\n\n\
% Handle internationalization\n\\usepackage[utf8]{inputenc}\n\\usepackage[T1]{fontenc}\n\
\\usepackage[right]{eurosym}\n\\usepackage[ngerman]{babel}\n\n% Handle graphics and tables\n\
\\usepackage{graphicx}\n\\usepackage{booktabs}\n\\usepackage{multirow}\n%\n% Handle hyperlinks\n\
\\usepackage[hidelinks]{hyperref}\n%\n% Use gray text for contact data\n\\usepackage{color}\n%\n\\usepackage{eurosym}\n"
#
# mehr Platz nach unten:
TXT_1 = TXT_1 + "%\n%\n\\setlength{\\textheight}{26cm}\n\\setlength{\\footskip}{0mm}\n\\setlength{\\footheight}{0mm}\n%\n"
#
# \\setkomavar{fromname}{RVE $\\cdot $ Habichtstra{\\ss}e 12 $\\cdot $ }\n\
#
# %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
#   RVE-Flag.png: 692 x 252 pixel
TXT_2 = "% Begin KOMA variables - make changes here\n\
\\newkomavar{gigdate}\n\
\\setkomavar{gigdate}{" + LSglobal.Datum + str(LSglobal.Jahr) + "}\n\
\\newkomavar{ausrichter}\n\
\\setkomavar{ausrichter}{Ruderverein Erlangen e.V.}\n\
\\setkomavar{fromname}{\\textbf{RVE}}\n\
\\setkomavar{fromaddress}{Habichtstra{\\ss}e 12\\\\91056 Erlangen}\n\
\\setkomavar{subject}{Rechnung:	Meldegebühr für die " + LSglobal.Zeit + "-Langstrecke }\n\
\\setkomavar{fromurl}{\\url{http://www.ruderverein-erlangen.de}}\n\
\\setkomavar{fromemail}{\\href{mailto:langstrecke@ruderverein-erlangen.de}{langstrecke@ruderverein-erlangen.de}}\n\
\\setkomavar{fromlogo}{\\includegraphics[height=1.76cm,width=2cm]{RVE-Flag.png}}\n\
\\setkomavar{yourref}[Ihre Teilnahme an der Langstrecke am ]{\\usekomavar{gigdate}}\n\
\n\
% Invoice data\n\
\\newkomavar{price}\n\
\\setkomavar{price}{\\EUR{105,00}}\n\
\\newkomavar{invoicedate}\n\
\\setkomavar{invoicedate}{FÄLLIGKEITSDATUM}\n\
\\newkomavar{iban}\n\
\\setkomavar{iban}{DE68 7635 0000 0000 0076 01}\n\
\\newkomavar{bic}\n\
\\setkomavar{bic}{BYLADEM1ERH}\n\n"

# %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
# \\footnotesize => \\scriptsize
TXT_3 = "% Begin header\n\
\\firsthead{\\null\\hfill \n\
 \\parbox[t][\\headheight][t]{10cm}{%\n\
  \\flushright\n\
  {\\small\\textsf{\n\
  \\color[gray]{.5}%\n\
  \\begin{tabular}[t]{r l}\n\
	\\textbf{\\usekomavar{ausrichter}} & \\multirow{5}{*}{\\usekomavar{fromlogo}} \\\\\n\
	Dr. Ulf Meerwald & \\\\\n\
	Regattaleitung &\\\\\n\
  \\usekomavar{fromaddress} & \\\\\n\
  \\end{tabular}\n\
  \\footnotesize{\\usekomavar{fromemail}}\\\\\n\
  \\footnotesize{\\usekomavar{fromurl}}\\\\\n\
  [\\baselineskip]\n\
  }}}%\n\
}\n\n\
% Begin footer\n\
\\firstfoot{\n\
\\parbox[t]{\\textwidth}{\n\
	\\scriptsize\n\
	\\color[gray]{.5}%\n\
	\\begin{tabular}[t]{l | r l | r l}\n\
		\\textbf{\\usekomavar{ausrichter}} &\n\
		\\multicolumn{2}{l|}{\\textbf{Kontakt}} &\n\
		\\multicolumn{2}{l}{\\textbf{Bankverbindung}}\\\\\n"

TXT_3 = TXT_3 + "		Habichtstra{\\ss}e 12 & \n\
		Web: & \\usekomavar{fromurl} &\n\
		IBAN: & \\usekomavar{iban}\\\\\n\
		91072 Erlangen &\n\
		E-Mail: & \\usekomavar{fromemail} &\n\
		BIC: & \\usekomavar{bic}\\\\\n\
	\\end{tabular}\n\
	}\n\
}\n\n"

TXT_4 = "% Begin main part - make changes here\n\
\\begin{document}\n\
\\begin{letter}{EMPFÄNGER\\\\\n\
	EMPFÄNGERADRESSE\\\\\n\
	12345 MUSTERSTADT}\n\n\
\\opening{Sehr geehrte Damen und Herren,}\n\n\
Für die Teilnahme an der Langstrecke vom Bayrischen Ruderverband und dem \\usekomavar{ausrichter} am \n\
\\usekomavar{gigdate} dürfen wir Ihnen folgendes berechnen:\n\n\\vspace{10pt}\n\n\
\\begin{tabular}{p{0.8\\textwidth} r}\n\
\\toprule\n\
\\textbf{Bezeichnung} & \\textbf{Betrag} \\\\\n\
\\midrule\n%\n"


####################################################################################################
# Vereine
#
Count_Boote   = 0
Count_Ruderer = 0
Count_Verein  = 0

Summe_Boote = 0
Summe_Kinder = 0

EUR_total = 0
sql = "SELECT * FROM verein WHERE kurz != 'RVE' AND dabei = 1"

# Gebühren für:
Kanal     = 18
Athletik  = 7
Bugnummer = 10
# Deckelung noch aus alten Zeiten (€ 250) jetzt auf übertriebene 2.500 EUR
Deckelung = 25000

fehlende_Bugnummern = [ 33 ]

Vcursor.execute(sql)
for Vsatz in Vcursor:
   # ------------------------------------- Kurzform mit Langform ersetzen
   Anzahl_Rennen = 0
   NoBoote = 0
   # ____________ Meldegeld
   EURO = 0
   # ____________ verspätete Abmeldung
   EURA = 0
   # ____________ verlorene Bugnummer
   EURB = 0
   # ____________ Nachmeldungen
   NACH = 0
   #
   TXT = TXT_1 + TXT_2 + TXT_3 
   TXT = TXT +  "% Begin main part - make changes here\n\\begin{document}\n\\begin{letter}{"
   TXT = TXT +  Vsatz[1] + "\\\\\n"
   TXT = TXT +  Vsatz[3] + "\\\\\n"
   TXT = TXT +  Vsatz[4] + "}\n%\n"
   TXT = TXT +  "\\opening{Sehr geehrte Damen und Herren,}\n\n"
   TXT = TXT +  "Für die Teilnahme an der Langstrecke vom Bayrischen Ruderverband und dem \\usekomavar{ausrichter} am \n"
   TXT = TXT +  "\\usekomavar{gigdate} dürfen wir Ihnen folgendes berechnen:\n\n\\vspace{10pt}\n\n"
   TXT = TXT +  "\\begin{tabular}{p{0.8\\textwidth} r}\n\\toprule\n\\textbf{Bezeichnung} & \\textbf{Betrag} \\\\\n"
   TXT = TXT +  "\\midrule\n%\n"
   #
   TXT_Fontsize = "\\footnotesize"
   TXTA = "Verspätete Abmeldungen: " +TXT_Fontsize + "{"
   #
   TXTB = "Verlorene Bugnummern: \\newline" +TXT_Fontsize + "{"
   #
   TXTM = "Meldegeld für folgende Mannschaften: " +TXT_Fontsize + "{\\newline\n%"
   #
   #___________________________ durchsuche Rennen
   sql = "SELECT * FROM rennen WHERE status >= 1"
   Rcursor.execute(sql)
   for Rsatz in Rcursor:
      Rennen       = Rsatz[0]
      RennenString = Rsatz[3]
      # 
      Ngray = 0
      Vrennen  = 0
      VArennen = 0
      sql = "SELECT * FROM boote  WHERE rennen = " + str(Rennen) + " ORDER BY startnummer"
      Bcursor.execute(sql)
      for Bsatz in Bcursor:
         StNr   = Bsatz[2]
         Boot   = Bsatz[0]
         Abmeldung = Bsatz[11]
         #
         #_____________________________________________ Höhe des Meldegeldes (über Strecke ermittelt)
         if(Rsatz[6] == "div."):
            Meldegeld = Athletik
         else:
            Meldegeld = Kanal
         #
         SH = ""
         Kommentar = Bsatz[13]
         if(Kommentar.find('achmeldung') > 0):
            Meldegeld = 2 * Meldegeld
            SH = "$^N$"
         #
         #
         # ===========================================================================================
         nPers   = 0
         sql = "SELECT * FROM r2boot  WHERE bootid = '" + Boot + "'"
         Qcursor.execute(sql)
         iR = 0
         for RudInd in Qcursor:
            sql = "SELECT * FROM ruderer WHERE id = '" + RudInd[2] + "'"
            Pcursor.execute(sql)
            Rd = Pcursor.fetchone()
            #
            if(Rd[7] == Vsatz[2]):
               nPers = nPers + 1
            if(iR == 0):
               # Name = "\\textbf{" + Rd[0] + " } " + Rd[1]
               Name =  Rd[1] + " " + Rd[2]
            else:
               # Name = Name + ", \\textbf{ " + Rd[0] + " } " + Rd[1]
               Name = "(" + Name + ", " + Rd[1] + " " + Rd[2] + ")"
            #               
            if(Vsatz[2] != Rd[7]):
               # Name = Name + " \\textcolor{gray}{\\scriptsize (" + Rd[6] + ")}"
               Name = Name + "$ ^{(" + Rd[7] + ")}$"
            #
            iR = iR + 1
         #
         if(nPers > 0):
            #
            if(Abmeldung == 0):
               NoBoote = NoBoote + 1
               if(iR == 2 and nPers == 1):
                  EURO = EURO + Meldegeld/2
                  Summe_Boote  = Summe_Boote + Meldegeld/2
               else:
                  EURO = EURO + Meldegeld
                  if(Rsatz[6] == "div."):
                     Summe_Kinder = Summe_Kinder + Meldegeld
                  else:
                     Summe_Boote  = Summe_Boote + Meldegeld
                  #---
               # TXT = TXT + VTXT
               Anzahl_Rennen = Anzahl_Rennen + 1
               #
               if(Vrennen == 0):
                  Vrennen = 1
                  TXTM = TXTM + "\n\\newline\\textbf{ Rennen " + str(Rennen) + ": " + RennenString + ":} "
                  TXTM = TXTM + Name + SH
               else:
                  TXTM = TXTM + ", " + Name + SH
               #
               if(SH == "$^N$"):
                  NACH = NACH + 1
                  print("=> Nachmeldung #" + str(Boot) + " - " + str(NACH)  + " (" + Vsatz[2] + ")")
               #
            # ===========================================================================================
            elif(Abmeldung > 1):  # verspätet abgemeldet
               if(VArennen == 0):
                  VArennen = 1
                  TXTA = TXTA + "\n\\newline\\textbf{Rennen " + str(Rennen) + ": " + RennenString + ": } "
                  TXTA = TXTA + "\\textcolor{red}{ " + Name + "}"
               else:
                  TXTA = TXTA + ", \\textcolor{red}{ " + Name + "}"
               #
               if(iR == 2 and nPers == 1):
                  EURA = EURA + Meldegeld/2
               else:
                  EURA = EURA + Meldegeld
               #____________________________ Nachmeldung
               if(SH == "$^N$"):
                  NACH = NACH + 1
                  print("=> Nachmeldung #" + str(Boot) + " - " + str(NACH) + " (" + Vsatz[2] + ") - verspätet abgemeldet")
             #
            if(StNr in fehlende_Bugnummern and Meldegeld == Kanal):
               TXTB = TXTB +  "	(" + str(StNr) + ") " + Name + "\\newline"
               EURB = EURB + 10
         #_______________________________________________________________________________________
         #
      #_______________________________________________________________________________________
   
   #_______________________________________________________________________________________
   #if(Vrennen > 0):
   #   TXT = TXT + "%\\n\\\\end{tabular}\\\\\\\\\\n%\\n%\\n"
   # Korrektur des 'ß' - sz:
   
   if(EURO > 0):
      euronen = "%6.2f"% (EURO)
      TXT = TXT + TXTM + "\n\\newline}	& " + euronen + " \\\\\n"
      TXT = TXT + "%\n% ... Meldegeld für " + str(NoBoote) + " Boote (genauer auf nächster Seite) ...\n%\n%\n"
      if(EURO > Deckelung):
         euronen = "%6.2f"% (EURO - Deckelung)
         TXT = TXT + "%\nDeckelung der Meldegebühren auf EUR " + str(Deckelung) + ": \\newline"
         TXT = TXT + " & \color{red}{-" + euronen + "} \\\\\n%\n"
         EURO = Deckelung
         
   if(EURA > 0):
      euronen = "%6.2f"% (EURA)
      TXT = TXT + TXTA +  "\\newline}	& " + euronen + " \\\\\n"
      # TXT = TXT + "%\nKulanz für die verspäteten Abmeldungen: \\newline"
      #TXT = TXT + "\\footnotesize\\textsf{Auf Grund der Corona-Pandemie waren Abmeldungen bis zum 22.10. um 14 Uhr kostenfrei. \n"
      #TXT = TXT + "Danach haben wir für das Boot Start- und Bugnummern eingetütet.\\newline} & \color{red}{" + euronen + "} \\\\\n%\n"
      # TXT = TXT + "%\nAufwandspauschale der entstandenen Kosten & 20,00 \\\\\n%"
      # EURA = 20
   #           
   if(EURB > 0):
      euronen = "%6.2f"% (EURB)
      TXT = TXT + TXTB +  "}\n	& " + euronen + " \\\\\n"
   TXT = TXT + "%\n\\midrule\n"
   #
   EUR_total = EUR_total + EURO + EURA + EURB
   #
   Count_Verein  = Count_Verein  + 1
   #                  #
   sql = "UPDATE verein SET rechnung = " + str(EURO + EURA + EURB) + " WHERE kurz = '" + Vsatz[2] + "' "
   cursor.execute(sql)
   connection.commit()
   #
   euronen = "%6.2f"% (EURO + EURA + EURB)
   TXT = TXT + "\\textbf{Gesamtbetrag (brutto):} & " + euronen + " \\\\\n"
   TXT = TXT + "\\bottomrule\n\\end{tabular}\n%\n\\vspace{10pt}\\\\\n%\n"
   #
   if(NACH == 1):
      TXT = TXT + "{\\scriptsize$^N$: Die Nachmeldung deutlich verspätet eingegangen, daher doppeltes Meldegeld.}\\\\\n"
   elif(NACH > 1):
      TXT = TXT + "{\\scriptsize$^N$: Nachmeldungen sind deutlich verspätet eingegangen, daher jeweils doppeltes Meldegeld.}\\\\\n"
   
   # TXT = TXT + "{\\scriptsize Die Rechnungsstellung erfolgt ohne Ausweis der Umsatzsteuer nach \\textsection19 UStG.\\vspace{5pt}\\\\}\n"
   TXT = TXT + "{\\scriptsize Die Rechnungsstellung erfolgt ohne Ausweis der Umsatzsteuer nach \\textsection19 UStG. (pro Boot 18\\euro, Athletiktest 7\\euro)\\vspace{5pt}\\\\}\n"
   #
   TXT = TXT + "Bitte benutzen Sie bei der Überweisung das Kennwort:\n\\\\ '\\textbf{Meldegebühr Langstreckentest " + Vsatz[2] + "'} \\vspace{1.2cm}\\\\\n"
   #TXT = TXT + "\\vspace{5pt}\\\\\nMit rudersportlichen Grü{\ss}en,\\vspace{-10pt}\\\\\n\\includegraphics[height=1.71cm,width=4.13cm]{UlfMeerwald.png}"
   TXT = TXT + "Vielen Dank im voraus, \\\\mit rudersportlichen Grü{\ss}en \\vspace{-2.1cm}\\\\ \\color{white}{.}\\hspace{7cm}\\color{white}{.}\n"
   TXT = TXT + "\\includegraphics[height=2.55cm,width=6.2cm]{UlfMeerwald.png}\n"
   #Vielen Dank im voraus, \vspace{0.2cm}\\mit rudersportlichen Grü{\ss}en \vspace{-2.1cm}\\ \color{white}{.}\hspace{7cm}\color{white}{.}
   TXT = TXT + "\n\\end{letter}\n\\end{document}\n"
   #
   TXT = TXT.replace('ß', '{\\ss}')
   # Dateiname ohne Leerzeichen:
   VtexName = Vsatz[2].replace(" ", "_")
   if(EURB > 0 and EURA > 0):
      print("pdflatex Rechnung_" + VtexName + ".tex : " + str(EURO) + " + " + str(EURA) + " spät abgemeldet + " + str(EURB) + " für Bugnr." )
   elif(EURB > 0):
      print("pdflatex Rechnung_" + VtexName + ".tex : " + str(EURO) + " + " + str(EURB) + " für Bugnr." )
   elif(EURA > 0):
      print("pdflatex Rechnung_" + VtexName + ".tex : " + str(EURO) + " + " + str(EURA) + " spät abgemeldet " )
   else:
      print("pdflatex Rechnung_" + VtexName + ".tex : " + str(EURO)  )
   fp = open("LaTeX/Rechnung_" + VtexName + ".tex","w")
   fp.write(TXT)
   fp.close()
   ws['A'+str(wsNr)] = Vsatz[1]
   ws['B'+str(wsNr)] = Vsatz[2]
   # ws['C'+str(wsNr)] = str(EURO + EURA + EURB)
   ws['C'+str(wsNr)] = (EURO + EURA + EURB)
   #
   sql = "SELECT * FROM betreuer WHERE verein = '" + Vsatz[2] + "' "
   Bcursor.execute(sql)
   Betreuer = " - "
   nB = 0
   for Bsatz in Bcursor:
      if(nB == 0):
         Betreuer = Bsatz[4]
      else:
        Betreuer = Betreuer +"\n" + Bsatz[4]
      nB = nB + 1
   #--------------
   ws['D'+str(wsNr)] = Betreuer
   #
   wsNr = wsNr + 1

#TXT = TXT_1 + TXT_2 + TXT_3 + TXT_4

##############################################################################################################
# Korrektur des 'ß' - sz:
#TXT = TXT.replace('ß', '{\\ss}')

# TXT = TXT + "\\n%======================\\n\\\\end{document}\\n"
connection.close()

# ______________________________________ save
wb.save('Rechnungen_' + LSglobal.Zeit + "_" + str(LSglobal.Jahr) + '.xlsx')

print("Gemeldet haben " + str(Count_Verein) + " Vereine und wir stellen " + str(EUR_total) + " EUR in Rechnung")
print("(Für gestartete Boote: " + str(Summe_Boote) + ", Bugnummern: " + str(Bugnummer * len(fehlende_Bugnummern)) + " EUR  und Athletik " + str(Summe_Kinder) + " EUR)")
