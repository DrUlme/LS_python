#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 15 19:51:37 2022

@author: ulf
"""

import os, sqlite3

#==============================================================================
# library für Excel:
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.formatting.rule import DataBarRule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.utils import range_boundaries
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import numbers
# image - using pillow?
from openpyxl.drawing.image import Image
# from PIL import Image


#==============================================================================
from termcolor import colored
#
# Text colors:
#   grey, red, green, yellow, blue, magenta, cyan, white
# Text highlights: 
#   on_grey,on_red, on_green, on_yellow, on_blue, on_magenta, on_cyan, on_white
# Attributes:
#   bold, dark, underline, blink, reverse, concealed
#========================================================================
# globale Parameter
import LSglobal

#========================================================================
# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect( LSglobal.SQLiteFile )


# Datensatzcursor erzeugen
cursor   = connection.cursor()
cursor_R = connection.cursor()
cursor_B = connection.cursor()
cursor_A = connection.cursor()
cursorRB = connection.cursor()

#==============================================================================

# globale Parameter


print(colored('Hello, World!', 'red', attrs=['underline','dark']) + " + " + colored('Hello, World!', 'red', attrs=['bold']))
print(colored('Hello, World!', 'green', attrs=['bold']))
print(colored('Hello, World!', 'yellow', attrs=['bold']))
print(colored('Hello, World!', 'red', attrs=['bold']))

# Erstellen eines Workbooks:
wb = Workbook()
ws = wb.active
ws.title = LSglobal.ZeitK+ str(LSglobal.Jahr)

# Schreibe Kopfzeile
ws.merge_cells('D1:J1')
ws['D1'] = "Lauf- und Athletiktest Frühjahr " + str(LSglobal.Jahr)
ws['D1'].font = Font(name='arial', sz=14, b=True, i=False, color='4444dd')

# SiegerLauf = 20
# SiegerTest = 5
# Absolviert = 4

# ws['O1'] = 'Sieger'
# ws['P1'] = SiegerLauf 
# ws['Q1'] = SiegerTest
# ws.merge_cells('R1:T1')
# ws['R1'] = 'Basis für Teilnahme: '
# ws['R1'].alignment = Alignment(horizontal="right")
# ws['U1'] = Absolviert

# --------------------------------------------
Zeile = 3
ws.merge_cells('B' + str(Zeile) +':B' + str(Zeile+1) )
ws.merge_cells('C' + str(Zeile) +':C' + str(Zeile+1) )
ws.merge_cells('D' + str(Zeile) +':D' + str(Zeile+1) )
ws.merge_cells('E' + str(Zeile) +':E' + str(Zeile+1) )
ws.merge_cells('I' + str(Zeile) +':L' + str(Zeile) )
ws.merge_cells('H' + str(Zeile) +':H' + str(Zeile+1) )
ws.merge_cells('N1:N' + str(Zeile) )

ws['A' + str(Zeile)] = "Start-"
ws['B' + str(Zeile)] = "Vorname"
ws['C' + str(Zeile)] = "Name"
ws['D' + str(Zeile)] = "Verein"
ws['E' + str(Zeile)] = "M/W"
ws['F' + str(Zeile)] = "Körperhöhe"
ws['G' + str(Zeile)] = "Gewicht"
ws['H' + str(Zeile)] = "Jahrgang"
ws['I' + str(Zeile)] = "Ergebnisse"
ws['I' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
ws['I' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
ws['I' + str(Zeile)].alignment = Alignment(horizontal="center")

ws['M' + str(Zeile)] = "Lauf"
ws['M' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
ws['M' + str(Zeile)].alignment = Alignment(horizontal="center")

ws.merge_cells('O1:U3')
ws['O1'] = 'Auswertung: Punkte'
ws['O1'].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')

# --------------------------------------------
Zeile = Zeile + 1


ws['A' + str(Zeile)] = "Nr."

ws['F' + str(Zeile)] = "[cm]"
ws['G' + str(Zeile)] = "[kg]"

# ws['H' + str(Zeile)] = "#1 Anzahl"
ws['I' + str(Zeile)] = "#1 Zeit"
ws['J' + str(Zeile)] = "#2 Anzahl"
ws['K' + str(Zeile)] = "#3 Anzahl"
ws['L' + str(Zeile)] = "#4 Anzahl"
ws['M' + str(Zeile)] = "[min:sec]"
ws['N' + str(Zeile)] = "Kommentar"
ws['N' + str(Zeile)].alignment = Alignment(horizontal="center")

# Vorname, Name, M/W, Körperhöhe, Körpergewicht, Jahrgang, Ergebnisse 1-4, Lauftest.
ws['O'+ str(Zeile)] = 'Gesamt'
ws['P'+ str(Zeile)] = 'Lauf'
ws['Q'+ str(Zeile)] = 'Athletik'
ws['R'+ str(Zeile)] = '#1'
ws['S'+ str(Zeile)] = '#2'
ws['T'+ str(Zeile)] = '#3'
ws['U'+ str(Zeile)] = '#4'


ws.column_dimensions["A"].alignment = Alignment(horizontal='center')
ws.column_dimensions["B"].alignment = Alignment(vertical='center')
ws.column_dimensions["C"].alignment = Alignment(vertical='center')
ws.column_dimensions["D"].alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions["E"].alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions["F"].alignment = Alignment(horizontal='center')

ws.column_dimensions["H"].alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions["I"].alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions["J"].alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions["K"].alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions["L"].alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions["M"].alignment = Alignment(horizontal='center', vertical='center')

CountMin = 10
###################################################################################################################
# Loop über Rennen
# --------------------------------------------
Zeile = Zeile + 1
# SQL-Abfrage
sql = "SELECT * FROM rennen WHERE nummer > 20"
# Empfang des Ergebnisses
cursor_R.execute(sql)
#
StartNr = 0
for dsatz in cursor_R:
   Rennen = dsatz[0]
   ReStr  = str(Rennen)
   # ----------------------------------------- hole Boote (= Athleten)
   sql = "SELECT * FROM boote WHERE rennen = " + str(Rennen) + " AND abgemeldet = 0"
   # Empfang des Ergebnisses
   cursor_B.execute(sql)
   Meldungen = len(cursor_B.fetchall())
   #
   Zeile = Zeile + 1
   # ----------------------------------------- Header pro Rennen
   ws['A' + str(Zeile)] = StartNr + 0.5
   ws['A' + str(Zeile)].font = Font(name='arial', sz=8, b=True, i=False, color='999999')
   ws['A' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   
   ws.merge_cells('B' + str(Zeile) +':C' + str(Zeile) )
   ws['B' + str(Zeile)]      = dsatz[1]
   ws['B' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['B' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   
   ws.merge_cells('D' + str(Zeile) +':H' + str(Zeile) )
   ws['D' + str(Zeile)]      = 'Bestes Ergebnis:'
   ws['D' + str(Zeile)].alignment = Alignment(horizontal='right', vertical='center')
   ws['D' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['D' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   #
   # Tests 1-4
   ws['I' + str(Zeile)]      = '=min(I' + str(Zeile+1) + ':I' + str(Zeile+Meldungen) + ')'
   ws['I' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['I' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   ws['I' + str(Zeile)].number_format = numbers.FORMAT_DATE_TIME5
   
   ws['J' + str(Zeile)]      = '=max(J' + str(Zeile+1) + ':J' + str(Zeile+Meldungen) + ')'
   ws['J' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['J' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   
   ws['K' + str(Zeile)]      = '=max(K' + str(Zeile+1) + ':K' + str(Zeile+Meldungen) + ')'
   ws['K' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['K' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   
   ws['L' + str(Zeile)]      = '=max(L' + str(Zeile+1) + ':L' + str(Zeile+Meldungen) + ')'
   ws['L' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['L' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   
   ws['M' + str(Zeile)]      = '=min(M' + str(Zeile+1) + ':M' + str(Zeile+Meldungen) + ')'
   ws['M' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['M' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   ws['M' + str(Zeile)].number_format = numbers.FORMAT_DATE_TIME5
   #
   # Tests 1-4
   # ws['Q' + str(Zeile)]      = '=IF(COUNT(H' + str(Zeile+1) + ':H' + str(Zeile+Meldungen) + ')> 0, AVERAGE(H' + str(Zeile+1) + ':H' + str(Zeile+Meldungen) + '), 0)'
   # ws['R' + str(Zeile)]      = '=IF(COUNT(I' + str(Zeile+1) + ':I' + str(Zeile+Meldungen) + ')> 0, AVERAGE(M' + str(Zeile+1) + ':M' + str(Zeile+Meldungen) + '), 0)'
   ws['R' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['R' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   ws['R' + str(Zeile)].number_format = numbers.FORMAT_DATE_TIME5
   
   #ws['S' + str(Zeile)]      = CountMin
   #'=IF(COUNT(I' + str(Zeile+1) + ':I' + str(Zeile+Meldungen) + ')> 0, AVERAGE(I' + str(Zeile+1) + ':I' + str(Zeile+Meldungen) + '), 0)'
   ws['S' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['S' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   
   #ws['T' + str(Zeile)]      = CountMin
   # '=IF(COUNT(J' + str(Zeile+1) + ':J' + str(Zeile+Meldungen) + ')> 0, AVERAGE(J' + str(Zeile+1) + ':J' + str(Zeile+Meldungen) + '), 0)'
   ws['T' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['T' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   
   #ws['U' + str(Zeile)]      = CountMin
   # '=IF(COUNT(K' + str(Zeile+1) + ':K' + str(Zeile+Meldungen) + ')> 0, AVERAGE(K' + str(Zeile+1) + ':K' + str(Zeile+Meldungen) + '), 0)'
   ws['U' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['U' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   
   # Lauf
   # ws['P' + str(Zeile)]      = '=IF(COUNT(M' + str(Zeile+1) + ':M' + str(Zeile+Meldungen) + ')> 0, AVERAGE(M' + str(Zeile+1) + ':M' + str(Zeile+Meldungen) + '), 0)'
   ws['P' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
   ws['P' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")   
   ws['P' + str(Zeile)].number_format = numbers.FORMAT_DATE_TIME5
   # 
   # Punktesumme
   ws['Q' + str(Zeile)] = '=max(Q' + str(Zeile+1) + ':Q' + str(Zeile+Meldungen)
   ws['Q' + str(Zeile)].font = Font(name='arial', sz=8, b=True, i=False, color='999999')
   ws['Q' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
   #
   ws['O' + str(Zeile)] = '=max(O' + str(Zeile+1) + ':O' + str(Zeile+Meldungen)
   ws['O' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='999999')
   ws['O' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")

   #
   
   # Create fill
   redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
   best_Fill = PatternFill(start_color='11FF11', end_color='11EE11', fill_type='solid')
   best_font = Font(size=14, bold=True, color="222222")
   
               
   #ws.conditional_formatting.add('A1:A10', formatting.CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=best_font))
   # ws.conditional_formatting.add('H' + str(Zeile+1) + ':H' + str(Zeile+Meldungen), CellIsRule(operator='between', formula=['10','$H'+str(Zeile)], stopIfTrue=True, fill=redFill))
   #  ws.conditional_formatting.add('K' + str(Zeile+1) + ':K' + str(Zeile+Meldungen), CellIsRule(operator='equal', formula=['$K$'+str(Zeile)], font=best_font ))
   # data_bar_rule = DataBarRule(start_type="num", start_value=1, end_type="num",  end_value='$J'+str(Zeile), color="0000FF00")  # Green
   # ws.conditional_formatting.add('I' + str(Zeile+1) + ':I' + str(Zeile+Meldungen), data_bar_rule)
   #
   #   ws.conditional_formatting.add('I' + str(Zeile+1) + ':I' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=1, end_type="num",  end_value='$I$'+str(Zeile), color="0000FF00"))
   ws.conditional_formatting.add('I' + str(Zeile+1) + ':I' + str(Zeile+Meldungen), CellIsRule(operator='equal', formula=['$I$'+str(Zeile)], font=best_font ) )
   #
   ws.conditional_formatting.add('J' + str(Zeile+1) + ':J' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=1, end_type="num",  end_value='$J$'+str(Zeile), color="0000FF00"))
   ws.conditional_formatting.add('J' + str(Zeile+1) + ':J' + str(Zeile+Meldungen), CellIsRule(operator='equal', formula=['$J$'+str(Zeile)], font=best_font ) )
   
   ws.conditional_formatting.add('K' + str(Zeile+1) + ':K' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=1, end_type="num",  end_value='$K$'+str(Zeile), color="0000FF00"))
   ws.conditional_formatting.add('K' + str(Zeile+1) + ':K' + str(Zeile+Meldungen), CellIsRule(operator='equal', formula=['$K$'+str(Zeile)], font=best_font ) )
   #
   ws.conditional_formatting.add('L' + str(Zeile+1) + ':L' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=1, end_type="num",  end_value='$L$'+str(Zeile), color="0000FF00"))
   ws.conditional_formatting.add('L' + str(Zeile+1) + ':L' + str(Zeile+Meldungen), CellIsRule(operator='equal', formula=['$L$'+str(Zeile)], font=best_font ) )
   # Lauf:
   ws.conditional_formatting.add('M' + str(Zeile+1) + ':M' + str(Zeile+Meldungen), CellIsRule(operator='equal', formula=['$M$'+str(Zeile)], stopIfTrue=True, font=best_font ) )
   
   # Lauf Auswertung:
   ws.conditional_formatting.add('P' + str(Zeile+1) + ':P' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=4, end_type="num",  end_value=30, color="00FF0000"))
   # Gesamt
   ws.conditional_formatting.add('O' + str(Zeile+1) + ':O' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=0, end_type="num",  end_value='$O$' + str(Zeile) , color="333333FF"))
   #
   ws.conditional_formatting.add('Q' + str(Zeile+1) + ':Q' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=0, end_type="num",  end_value='$Q$' + str(Zeile) , color="0000AA00"))
   #
   ws.conditional_formatting.add('R' + str(Zeile+1) + ':R' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=0, end_type="num",  end_value=15, color="0000FF00"))
   ws.conditional_formatting.add('S' + str(Zeile+1) + ':S' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=0, end_type="num",  end_value=15, color="0000FF00"))
   ws.conditional_formatting.add('T' + str(Zeile+1) + ':T' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=0, end_type="num",  end_value=15, color="0000FF00"))
   ws.conditional_formatting.add('U' + str(Zeile+1) + ':U' + str(Zeile+Meldungen), DataBarRule(start_type="num", start_value=0, end_type="num",  end_value=15, color="0000FF00"))
   #
   #zZ
   # new_range = openpyxl.workbook.defined_name.DefinedName('Boote_' + ReStr, attr_text='Rennen!$C$' + str(zeile))
   # wb.defined_names.append(new_range)
   myEven = 0
   # Summe der Meldungen pro Rennen (ändert sich mit Änderung auf der Hauptseite
   # ws2['C' + str(zeile)] = "=(SUMIF('Meldungen'!$A$7:$A$256,$A" + str(zeile) + " ) - $A"  + str(zeile) + ") / $A"  + str(zeile)
   cursor_B.execute(sql)
   Zeile0 = Zeile
   ZS = str(Zeile + 1)
   ZE = str(Zeile + Meldungen)
   for ds in cursor_B:
      Zeile = Zeile + 1
      sZ = str(Zeile)
      StartNr = StartNr + 1
      # suche nach Einträgen in r2boot for dieses Boot
      sql = "SELECT * FROM r2boot WHERE bootNr = " + str(ds[0])
      cursorRB.execute(sql)
      # 
      rbs = cursorRB.fetchone()
      # suche nach Einträgen in r2boot for dieses Boot
      sql = "SELECT * FROM ruderer WHERE nummer = " + str(rbs[3])
      cursor_A.execute(sql)
      Rd = cursor_A.fetchone()
      Vorname  = Rd[1]
      Name     = Rd[2]
      MW       = Rd[3]
      Jahrgang = Rd[4]
      #
      print('Boot #' + str(ds[0]) + ': r2boot ' + str(rbs[0]) + ' : #' + str(rbs[3]) + ' ' + Vorname + ' ' + Name)
      #
      # setze Zeitformat
      ws['I' + str(Zeile)].number_format = numbers.FORMAT_DATE_TIME5
      ws['M' + str(Zeile)].number_format = numbers.FORMAT_DATE_TIME5
      #
      # startnummer bereits gesetzt?
      #sonst setzen und auch in sql-Datenbank setzen
      ws['A' + str(Zeile)] = StartNr
      ws['B' + str(Zeile)] = Vorname
      ws['C' + str(Zeile)] = Name
      ws['D' + str(Zeile)] = Rd[7]
      ws['H' + str(Zeile)] = Jahrgang
      ws['E' + str(Zeile)] = MW
      # Kommentar aus Boot hinzufügen:
      ws['N' + str(Zeile)] = ds[11]
      
      # ============================================================================================ Auswertung
      # Platzierung	Wertung Athletikübungen	Wertung Lauftest
      #  #2	#3	#4	 
      # 1	15	15	15	15	30
      # 2	12	12	12	12	24
      # 3	10	10	10	10	20
      # 4	9	9	9	9	18
      # 5	8	8	8	8	16
      # 6	7	7	7	7	14
      # 7	6	6	6	6	12
      # 8	5	5	5	5	10
      # 9	4	4	4	4	8
      # ≥10	3	3	3	3	6
      ws['O' + str(Zeile)]      = '=P' + str(Zeile) + '+Q' + str(Zeile) + ''
      ws['Q' + str(Zeile)]      = '=SUM(R' + str(Zeile) + ':U' + str(Zeile) + ')'
      # 1 : Parcour
      ws['R' + sZ]      = '=IF(I' + sZ + '> 0' + ',IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ',1) = 1,15,IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ', 1) = 2,12,' + \
         'IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ',1) = 3,10,IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ', 1) = 4,9,' + \
         'IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ',1) = 5,8,IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ', 1) = 6,7,' + \
         'IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ',1) = 7,6,IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ', 1) = 8,5,' + \
         'IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ',1) = 9,4,IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ', 1) = 10,3,' + \
         'IF(rank(I' + sZ + ',I' + ZS + ':I' + ZE + ',1) = 11, 2, 1))))))))))),0)'
      # ws['Q' + str(Zeile)]      = '=IF(H' + str(Zeile) + '>$Q$' + str(Zeile0) + ',round((H' + str(Zeile) + '-$Q$' + str(Zeile0) + ')/($H$' + str(Zeile0) + '-$Q$' + str(Zeile0) + ')*$P$1)+$T$1,IF(H' + str(Zeile) + '>0, $T$1, 0))'
      #
      # 2.
      ws['S' + sZ]      = '=IF(J' + sZ + '>$S$' + str(Zeile0) + ',IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ',0) = 1,15,IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ', 0) = 2,12,' + \
         'IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ',0) = 3,10,IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ', 0) = 4,9,' + \
         'IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ',0) = 5,8,IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ', 0) = 6,7,' + \
         'IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ',0) = 7,6,IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ', 0) = 8,5,' + \
         'IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ',0) = 9,4,IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ', 0) = 10,3,' + \
         'IF(rank(J' + sZ + ',J' + ZS + ':J' + ZE + ',0) = 11, 2, 1))))))))))),0)'
      # 3.
      ws['T' + sZ]      = '=IF(K' + sZ + '>$T$' + str(Zeile0) + ',IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ',0) = 1,15,IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ', 0) = 2,12,' + \
         'IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ',0) = 3,10,IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ', 0) = 4,9,' + \
         'IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ',0) = 5,8,IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ', 0) = 6,7,' + \
         'IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ',0) = 7,6,IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ', 0) = 8,5,' + \
         'IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ',0) = 9,4,IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ', 0) = 10,3,' + \
         'IF(rank(K' + sZ + ',K' + ZS + ':K' + ZE + ',0) = 11, 2, 1))))))))))),0)'
      #ws['R' + str(Zeile)]      = '=IF(I' + str(Zeile) + '>$R$' + str(Zeile0) + ',round((I' + str(Zeile) + '-$R$' + str(Zeile0) + ')/($I$' + str(Zeile0) + '-$R$' + str(Zeile0) + ')*$P$1)+$T$1,IF(I' + str(Zeile) + '>0, $T$1, 0))'
      #ws['S' + str(Zeile)]      = '=IF(J' + str(Zeile) + '>$S$' + str(Zeile0) + ',round((J' + str(Zeile) + '-$S$' + str(Zeile0) + ')/($J$' + str(Zeile0) + '-$S$' + str(Zeile0) + ')*$P$1)+$T$1,IF(J' + str(Zeile) + '>0, $T$1, 0))'
      # ws['T' + str(Zeile)]      = '=IF(K' + str(Zeile) + '>$T$' + str(Zeile0) + ',round((K' + str(Zeile) + '-$T$' + str(Zeile0) + ')/($K$' + str(Zeile0) + '-$T$' + str(Zeile0) + ')*$P$1)+$T$1,IF(K' + str(Zeile) + '>0, $T$1, 0))'
      ws['U' + sZ]      = '=IF(L' + sZ + '>$U$' + str(Zeile0) + ',IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ',0) = 1,15,IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ', 0) = 2,12,' + \
         'IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ',0) = 3,10,IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ', 0) = 4,9,' + \
         'IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ',0) = 5,8,IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ', 0) = 6,7,' + \
         'IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ',0) = 7,6,IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ', 0) = 8,5,' + \
         'IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ',0) = 9,4,IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ', 0) = 10,3,' + \
         'IF(rank(L' + sZ + ',L' + ZS + ':L' + ZE + ',0) = 11,2,1))))))))))),0)'
      # =WENN(K7>$T$6;WENN(RANG(K7;K7:K8; 0) = 1; 15;WENN(RANG(K7;K7:K8; 0) = 2; 12; 3)); 0)
      #
      #Lauf:
      ws['P' + sZ]      = '=IF(M' + sZ + '>0' + ',IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 1,30,IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ', 1) = 2,24,' + \
         'IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 3,20,IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ', 1) = 4,18,' + \
         'IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 5,16,IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ', 1) = 6,14,' + \
         'IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 7,12,IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ', 1) = 8,10,' + \
         'IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 9,8,IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ', 1) = 10,6,' + \
         'IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 11, 4, 2))))))))))),0)'
      # ws['P' + sZ]      = '=IF(M' + sZ + '<$P$' + str(Zeile0) + ',IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 1,15,IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ', 1) = 2,12,' + \
      #    'IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 3,10,IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ', 1) = 4,9,' + \
      #    'IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 5,8,IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ', 1) = 6,7,' + \
      #    'IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 7,6,IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ', 1) = 8,5,' + \
      #    'IF(rank(M' + sZ + ',M' + ZS + ':M' + ZE + ',1) = 9, 4, 3))))))))),0)'
      #  # ws['O' + str(Zeile)]      = '=IF(L' + str(Zeile) + '<$O$' + str(Zeile0) + ',round((L' + str(Zeile) + '-$O$' + str(Zeile0) + ')/($L$' + str(Zeile0) + '-$O$' + str(Zeile0) + ')*$O$1)+$T$1,IF(L' + str(Zeile) + '>0, $T$1, 0))'
      #
      ws['O' + str(Zeile)].font = Font(name='arial', sz=13, b=True, i=False, color='4444dd')
      ws['Q' + str(Zeile)].font = Font(name='arial', sz=12, b=True, i=False, color='4444dd')
      #
      # if(myEven == 1):
      #     ws['A' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['B' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['C' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['D' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['E' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['F' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['G' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['H' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['I' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['J' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['K' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['L' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     ws['M' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")
      #     myEven = 0
      # else:
      #     myEven = 1
      #     ws['A' + str(Zeile)].fill = PatternFill(start_color="dddddd", end_color="dddddd",  fill_type = "solid")



###################################################################################################################
# setze Ausrichtung für Telefon-Nummer auf links zurück
# ws['E8'].alignment = Alignment(horizontal="center")
# erstelle Filter
# maxCols = str( zeile + 10 ) auf 256 gesetzt

# fixiere Tabelle:
ws.freeze_panes = ws['D6']

ws.auto_filter.ref = "A5:U256"


logo = Image("RVE_BRV_Flag.png")

# A bit of resizing to not fill the whole spreadsheet with the logo
logo.height = 77
logo.width = 210

# ws.add_image(logo, "I1")
ws.add_image(logo, "N1")

# Breite festlegen
ws.column_dimensions['A'].width = "6"  # Nr
ws.column_dimensions['B'].width = "16" # Vorname
ws.column_dimensions['C'].width = "16" # Name
ws.column_dimensions['D'].width = "16" # Verein
ws.column_dimensions['E'].width = "5"  # M/W
ws.column_dimensions['F'].width = "8"  # Körperhöhe
ws.column_dimensions['G'].width = "8"  # Gewicht
ws.column_dimensions['H'].width = "8"  # Jahrgang
ws.column_dimensions['I'].width = "10"  #1
ws.column_dimensions['J'].width = "10" #2
ws.column_dimensions['K'].width = "10" #3
ws.column_dimensions['L'].width = "10" #4
ws.column_dimensions['M'].width = "16" # Lauf
ws.column_dimensions['N'].width = "28"  # Kommentar / Bild


wb.save('Lauf-Athletiktest_' + LSglobal.ZeitK + "_" + str(LSglobal.Jahr) + '.xlsx')
