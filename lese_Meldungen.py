#  python3
#
# needs OpenPYXL:
#   /usr/bin/python3 -c "import openpyxl; print(openpyxl)" 
# 
# has issues with SoftMaker Office !!! 
# XML support:
# from lxml import etree
import os, sys, sqlite3
import time
from termcolor import colored

# Excel
#========================================================================
from openpyxl import load_workbook

# globale Parameter
import LSglobal

Pfad_LS        = os.getcwd()
print("Starte aus Pfad '" + Pfad_LS + "'")

#========================================================================
# Verbindung zur Datenbank erzeugen
connection = sqlite3.connect( LSglobal.SQLiteFile )

# Datensatzcursor erzeugen
cursor  = connection.cursor()
bcursor = connection.cursor()
rcursor = connection.cursor()
r2bCrs  = connection.cursor()

# hole Renn-Nummern für Früh und Spät-Starter
sql = "SELECT wert FROM meta WHERE name = 'Frühstarter'"
cursor.execute(sql)
Rd = cursor.fetchone()
Frühstart = Rd[0]
iFrüh = int(Frühstart)

sql = "SELECT wert FROM meta WHERE name = 'Spätstarter'"
cursor.execute(sql)
Rd = cursor.fetchone()
Spätstart = Rd[0]
iSpät = int(Spätstart)

print("__________________________________________________________\n")
Pfad_Meldungen = Pfad_LS + '/' + LSglobal.MeldeDir
print("Suche Files in Pfad '" + Pfad_Meldungen + "'")

#os.chdir( Pfad_Meldungen )
#FILES=os.listdir( '.' )
# FILES=os.listdir( Pfad_Meldungen )
# Get list of all files in a given directory sorted by name
FILES = sorted( filter( lambda x: os.path.isfile(os.path.join(Pfad_Meldungen, x)),
                        os.listdir(Pfad_Meldungen) ) )
for filename in FILES:
   print(filename)
   nFN = len(filename)
   Endung = filename[nFN-4:nFN]
   if(Endung.find("xlsx")>=0):
      print("__________________________________________________________\n")
      # get change time
      
      # open existent workbook:
      print("_________________________________________________________ " + filename )
      print(Pfad_Meldungen + "/" + filename)
      wb = load_workbook(Pfad_Meldungen + "/" + filename)
      # wb = load_workbook(filename = '../F2020/Meldungen/01__RVE_Thea.xlsx')
      # create a new Workbook:
      # from openpyxl import Workbook
      print(wb.sheetnames)
      ws = wb.active
      
      
      #========================================================================
      Name_Rennen_7 = str.strip(ws['L8'].value)
      if(Name_Rennen_7 == "JM 2- B"):
         print("Fehlerhaftes Meldeformular...")
         mitFehler = 1
      else:
         mitFehler = 0
      # exit()
      #========================================================================
      Vereinsname = str.strip(ws['A3'].value)
      Verein      = str.strip(ws['G3'].value)
      
      Adresse_1 =  ws['A5'].value
      Adresse_2 =  ws['A6'].value
      if( Adresse_2 == None): 
         Adresse_2 = ""
      if(Verein == None):
         print(filename + " hat kein Vereinskürzel, breche ab...")
         exit()
      sql = "SELECT EXISTS(SELECT 1 FROM verein WHERE kurz='" + Verein + "' LIMIT 1)"
      cursor.execute(sql)
      record = cursor.fetchone()
      
      if record[0] == 1:
         print (Verein +  " ist schon in der Datenbank!" )
         sql = "UPDATE verein SET dabei = 1 WHERE kurz='" + Verein + "' "
         cursor.execute(sql)
         connection.commit()
      else:
         sql = "INSERT INTO verein VALUES(" \
            "'unknown ID of " + Verein + "', " \
            "'" + Vereinsname + "', '" + Verein + "', '" \
            + Adresse_1 + "', '" + Adresse_2 + "', " \
            "0.0, 1, 1)"
         # print(sql)
         cursor.execute(sql)
      
      sqlite_select_query = """SELECT count(*) from verein"""
      cursor.execute(sqlite_select_query)
      totalRows = cursor.fetchone()
      print("Total rows are:  ", totalRows[0])
      
      print(' ____________________________________ ')
      print(' ')
      
      # ========================================================================
      
      Vorname  = str.strip(ws['A8'].value)
      Name     = str.strip(ws['C8'].value)
      Phone    = ws['E8'].value
      email    = ws['H8'].value
      
      #   "name" TEXT, "vorname" TEXT, "telefon" TEXT, "verein" TEXT, "email" TEXT
      # And this is the named style:
      # cursor.execute("select * from betreuer where vorname=:Vorname and name=:Name and verein=:Verein ", {"Vorname": Vorname, "Name": Name, "Verein": Verein})
      # print( cursor.fetchone())
      
      #
      if(Name == None):
         Hname = Vorname.split()
         Hlen = len(Hname)
         Vorname = Hname[0]
         Name = Hname[Hlen-1]
         
      sql = "SELECT EXISTS(SELECT 1 FROM betreuer WHERE verein='" + Verein + "' and name='" + Name + "' LIMIT 1)"
      cursor.execute(sql)
      record = cursor.fetchone()
      
      if record[0] == 1:
         print ("Betreuer: " + Vorname + " " + Name + " (" + Verein + ") ist schon in der Datenbank!" )
      else:
         sql = "INSERT INTO betreuer VALUES( " \
            "'" + Vorname + "', '" + Name + "', '" + Verein + "', " \
            "'" + Phone + "', '" + email + "' )"
         cursor.execute(sql)
         connection.commit()
         print(sql)
      
      print(' ____________________________________ ')
      print(' ')
      
      # ========================================================================
      
      sqlite_select_query = """SELECT count(*) from ruderer"""
      cursor.execute(sqlite_select_query)
      tmp = cursor.fetchone()
      nRuderer = tmp[0]
      print("Anzahl der Ruderer in Datenbank:  ", nRuderer)
      
      sqlite_select_query = """SELECT count(*) from boote"""
      cursor.execute(sqlite_select_query)
      tmp = cursor.fetchone()
      nBoote = tmp[0]
      print("Anzahl der Boote in Datenbank:  ", nBoote)
      
      sqlite_select_query = """SELECT count(*) from r2boot"""
      cursor.execute(sqlite_select_query)
      tmp = cursor.fetchone()
      nR2Boot = tmp[0]
      print("Anzahl der Ruderer in Booten in der Datenbank:  ", nR2Boot)
      
      # indices für Ruderer (max. 4+)
      Ruderer = [None] * 5
      # =======================================================================   Loop über die Zeilen der Meldung   ==============
      Position = 11
      while Position > 7:
         Names = ","
         NR = 0
         BootNr = -1
         # ____________________________________________ Rennen und Bootsgattung
         Rennen   = ws['B' + str(Position)].value
         if(Rennen == None):
            Rennen = 0;
         print("Starte mit Eintrag für Rennen " + str(Rennen) + " ..." )
         if Rennen > 0:
            Comment  = ws['I' + str(Position)].value
            if( Comment == None):
               Comment = "-"
            # ======================================================================== Fehler im Formular !!!
            if(mitFehler == 1):
               if(Rennen == 7):
                  print("Junioren B 2- gemeldet ! Ist aber nicht erlaubt.")
                  exit()
               elif(Rennen > 7 and Rennen <= iSpät):
                  Rennen = Rennen - 1
            # ======================================================================== Fehler im Formular ====
            #
            sql = "SELECT * FROM rennen WHERE nummer='" + str(Rennen) + "' "
            cursor.execute(sql)
            record = cursor.fetchone()
            #______________________________ string für 6000 m -Auswertung
            if(record[6] == "6000 m"):
               String6000 = "0"
            else:
               String6000 = "-"
               # print(record[6])
            #---
            AlterStart = record[10]
            AlterEnd   = record[11]
            # print(record)
            # ToDo: check für 3 Rennen ohne Zuordnung
            Gender = record[4]
            # print(record[7])
            if(record[9] > 0):
               LGWi = 1
               Gewicht = -1.0
            else:
               anz = Comment.lower().find('lgw') + Comment.lower().find('leicht')+2
               if(anz > 0):
                  LGWi = 1
                  Gewicht = -1.0
               else:
                  LGWi = 0
                  Gewicht = 0.0
               # anz = Comment.find('LGW') + Comment.find('Lgw') + Comment.find('eichtgewicht')
               # print(anz)
            #
            #____________________________________ Boot -> Anzahl Ruderer (NR)
            Boot  = record[5]
            if Boot == 'all':
               Boot = ws['H' + str(Position)].value
            #
            if   Boot == '1x':
               NR = 1
            elif Boot == '2x':
               NR = 2
            elif Boot == '2-':
               NR = 2
            elif Boot == '4x':
               NR = 4
            elif Boot == '4-':
               NR = 4
            elif Boot == '4x+':
               NR = 5
            elif Boot == '4+':
               NR = 5
            elif Boot == 'Athletik':
               NR = 1
            else:
               NR = 0
            #____________________________________ richtiges Rennen / mit Bootsangabe
            if(NR == 0):
              print( "in '" + filename + "' ist in Zeile " + str(Position) + " die Bootsgattung falsch")
              exit()
            if(record[5] == 'All' and Rennen > 1):
              print( "in '" + filename + "' ist in Zeile " + str(Position) + " Früh-/Spätstarter eingetragen !")
              exit()
            #
            BootID = '0'
            
            #Gender   = ws['E' + str(Position + iP)].value
            #LGW      = ws['D' + str(Position + iP)].value
            
            # Ruderer überprüfen: in Datenbank => suche rudererid in r2boot => bootid => 
            
            # =================================================================   Loop über die Ruderer   ==============
            # Liste für die Ruderer-ID's
            myIDs = []
            Platz = []
            #
            Alternativ = '0';
            for iP in range(NR):
              Vorname  = str.strip(ws['C' + str(Position + iP)].value)
              Name     = str.strip(ws['D' + str(Position + iP)].value)
              Jahr     = ws['E' + str(Position + iP)].value
              Verein2  = ws['F' + str(Position + iP)].value
              #
              iComment = ws['I' + str(Position + iP)].value
              #
              # Test der Zeile für den Ruderer
              if( (Vorname == "Rudi" and Name == "Riemen" ) or (Vorname == "Sabine" and Name == "Skull")):
                 print("File '" + filename + "' noch mit Dummy-Namen ! bitte bearbeiten !")
                 sys.exit()
              if( Vorname == None): 
                print( "in '" + filename + "' fehlt in Zeile " + str(Position) + " der Vorname")
                exit()
              if( Name == None): 
                print( "in '" + filename + "' fehlt in Zeile " + str(Position) + " der Name")
                exit()
              #--  Check Alter
              if( Jahr == None): 
                 print( "in '" + filename + "' fehlt in Zeile " + str(Position) + " das Jahr")
                 exit()
              if(Jahr < AlterStart ):
                 print( "in '" + filename + "' ist der Ruderer " + Vorname + " " + Name + " in Zeile " + str(Position) + " zu Alt")
                 exit()
              if(Jahr > AlterEnd ):
                 print(colored('WARNING: ', 'red', attrs=['bold']), "Ruderer " + Vorname + " " + Name + " in Zeile " + str(Position) + " ist zu jung !!!" + colored(' -ALTER-!!!', 'red', attrs=['bold']))
              #
              if( iComment == None): 
                 iComment = '-'
              # find 'früh' und 'spät' in Kommentar und setze die alternative Variable
              if(iComment.lower().find('früh') > 0):
                 Alternativ = Frühstart
              elif(iComment.lower().find('spät') > 0):
                 Alternativ = Spätstart
              #---
              if(iComment.lower().find('stm.') > 0):
                 Platz.append(-1)
              else:
                 Platz.append(iP + 1)
              #---
              sql = "SELECT id FROM ruderer WHERE name='" + Name + \
              "' and vorname='" + Vorname + "' " + " and jahrgang=" + str(Jahr) + " LIMIT 1"
              cursor.execute(sql)
              rCheck = cursor.fetchone()
              # check if is in list
              if(rCheck != None):
                 myIDs.append( rCheck[0] )
                 # print('sql in r2boot...')
                 sql = "SELECT bootid FROM r2boot WHERE rudererid='" + rCheck[0] + "' "
                 # print(sql)
                 r2bCrs.execute(sql)
                 r2b = r2bCrs.fetchone()
                 if(r2b != None):
                    # print('sql in boote ... ' + r2b[0])
                    sql = "SELECT rennen FROM boote WHERE id='" + r2b[0] + "' "
                    # print(sql)
                    bcursor.execute(sql)
                    gemRennen = bcursor.fetchone()
                    if(gemRennen != None):
                       # print('gleiches Rennen ? - ' + str(gemRennen[0]))
                       if(gemRennen[0] == Rennen):
                          BootID = r2b[0]
                          # print('Ruderer ' + Vorname + ' ' + Name + ' ist auch in Rennen ' + str(gemRennen[0]) + ' gemeldet')
                          print(colored('WARNING: ', 'red', attrs=['bold']), "Ruderer ist bereits in Boot " + str(BootID) + " gemeldet")
                       else:
                          print(colored('WARNING: ', 'red', attrs=['bold']), "Ruderer ist auch in Rennen " + str(gemRennen[0]) + " gemeldet")
                 else:
                    print ("Ruderer: " + Vorname + " " + Name + " (" + str(Jahr) + ", " + \
                      Verein + ") ist schon in der Datenbank! (ohne Rennen)" + str(Ruderer[iP]) )

              #-------------------- if Ruderer not exist:
              else:
                  # ???
                  if(iP == 0):
                     JahrBoot = Jahr
                  elif(JahrBoot > Jahr):
                     JahrBoot = Jahr
                  #______________________
                  if Gender == "w":
                      Gender = "F"
                  elif Gender == "W":
                      Gender = "F"
                  elif Gender == "f":
                      Gender = "F"
                  elif Gender == "m":
                      Gender = "M"
                  #______________________
                  if( Verein2 == None): 
                     Verein2 = Verein
                  #
                  nRuderer = nRuderer + 1
                  Names = Names + str(nRuderer) + ","
                  #
                  sql = "INSERT INTO ruderer VALUES( " + str(nRuderer) \
                     + ", '" + Vorname + "', '" + Name + "', '" + Gender + "', " + str(Jahr) + ", " \
                    + str(LGWi) + ", -1.0, '" + Verein2 + "', '" + Alternativ + "', -1 )"
                  print(sql)
                  
                  cursor.execute(sql)
                  connection.commit()
                  myIDs.append( str(nRuderer) )
                  #
                  # definiere den Ruderer-index für den Ruderer auf Platz iP
                  Ruderer[iP] = nRuderer
              #--------------------  Ruderer ist in Datenbank / neu hinzugefügt
            #-- Ende Schleife Ruderer
            #
            # print("Boot: '" + Boot + "': " + Names + " (" + str(JahrBoot) + ") in Rennen " + str(Rennen))
            
            if(BootID == '0'):
               # =================================================================   Wenn noch nicht gemeldet
               nBoote = nBoote + 1
               bootID = str(nBoote)
               #
               sql = "INSERT INTO boote VALUES( '" \
                  + bootID + "', '0', 0, " + str(Rennen) + ", " \
                  + "'0',   '0', '0', '" + String6000 + "',   '0', '" + String6000 + "',   '0',   0, " \
                  + "'" + Alternativ + "', '" + Comment + "')"
               # ToDo: Früh / Spät vor Comment einsetzen (alternativ) => manuell
               cursor.execute(sql)
               connection.commit()
               # ______________________ Füge Ruderer zu r2boot-Liste
               for iP in range(NR):
                   #iP in myIDs:
                   # print(iP)
                   # sql = "UPDATE ruderer SET boot = " + str(nBoote) + " WHERE nummer = " + str(Ruderer[1 + iP])
                   nR2Boot = nR2Boot + 1   # Melde-ID
                   sql = "INSERT INTO r2boot VALUES( '" \
                     + str(nR2Boot) + "', '"  + bootID + "', '"  \
                     + str(myIDs[iP]) + "', " + str(Platz[iP]) + " )"
                   # old: + str(nR2Boot) + ", " + str(Rennen) + ", " + str(nBoote) + ", "  \
                   #      + str(Ruderer[iP]) + ", " + str(iP) + " )"
                   # r2boot:  "nummer, rennNr, bootNr, verein, rudererid, platz INTEGER"
                   print(sql)
                   cursor.execute(sql)
                   connection.commit()
            #---------
            #      print("Rennen: '" + Rennen + "': " + Vorname + " " + Name + " (" + str(Jahr) + ")" + Comment)
            #print(ws[cellA].value + ' ' + ws['F' + str(Position)].value + ' ' + ws['G' + str(Position)].value + ' (' \
            #+ ws['E' + str(Position)].value + ') ' + str(ws['H' + str(Position)].value) )
         print(' --- ')
         if Position > 50:
            Position = 0
         elif NR == 0:
            Position = 0
         else:
            Position = Position + NR
            
      #---
      #sheet_ranges = wb['range names']
      #print(sheet_ranges['A1'].value)
      # ==============================================================================================================
      wb.close()
      print("=========================================================================================================\n")
   else:
      print("'" + Endung + "' ist kein Excel-Format !")
      print("=========================================================================================================\n")
#-----

# Suche Rennen nach

connection.close()
# import xlrd
#
# book = xlrd.open_workbook('../Meldeliste_RVE.xlsx')
# sheet = book.sheet_by_index(1)
# cell = sheet.cell(1,1)
# print(cell)
os.system( 'python LS_update_DB.py')


